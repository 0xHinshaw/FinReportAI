from logger import Logger
from datetime import datetime
from collections import defaultdict
import signal
import re
from concurrent.futures import ProcessPoolExecutor
import glob
from pdfminer.converter import PDFPageAggregator
from matplotlib.patches import Rectangle
from matplotlib.lines import Line2D
from pdfminer.layout import LAParams, LTTextBox, LTTextLine, LTImage, LTLine, LTRect, LTChar, LTTextBoxHorizontal,LTTextContainer
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.pdfpage import PDFPage
from pdfminer.pdfparser import PDFParser
from pdfminer.pdfdocument import PDFDocument
import matplotlib.pyplot as plt
from openpyxl import Workbook
from math import sqrt
import openpyxl
from cell_class import excel_cell
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import os
from tqdm import tqdm
from collections import defaultdict

logger = Logger(log_file_path=f'/data/shenyuge/lingyue-data-process/pdflux/log{datetime.now().strftime("%Y-%m-%d %H:%M")}.log')

class UnionFind:
    def __init__(self, size):
        self.parent = list(range(size))
        self.rank = [1] * size

    def find(self, p):
        if self.parent[p] != p:
            self.parent[p] = self.find(self.parent[p])  # 路径压缩
        return self.parent[p]

    def union(self, p, q):
        rootP = self.find(p)
        rootQ = self.find(q)
        if rootP != rootQ:
            if self.rank[rootP] > self.rank[rootQ]:
                self.parent[rootQ] = rootP
            elif self.rank[rootP] < self.rank[rootQ]:
                self.parent[rootP] = rootQ
            else:
                self.parent[rootQ] = rootP
                self.rank[rootP] += 1

    def connected(self, p, q):
        return self.find(p) == self.find(q)

def group_intersections(intersections, num_objects):
    uf = UnionFind(num_objects)  # 创建并查集，大小为 num_objects，假设最大编号为 num_objects-1
    for i, j in intersections:
        uf.union(i, j)
    
    groups = {}
    for object_id in range(num_objects):
        root = uf.find(object_id)
        if root not in groups:
            groups[root] = []
        groups[root].append(object_id)
    
    return list(groups.values())

def modify_bbox(bbox):
    for index,item in enumerate(bbox):
        bbox[index] = round(item)
    return bbox

def create_index_to_column_dict(n):
    """
    创建一个从整数索引到Excel列字母的映射字典。
    参数:
        n (int): 最大索引数，例如，如果n=28，将生成从1到28的映射。
    返回:
        dict: 映射字典，键为整数，值为对应的Excel列字母。
    """
    return {i-1: get_column_letter(i) for i in range(1, n + 1)}

def distance(point1, point2):
    return sqrt((point1[0] - point2[0]) ** 2 + (point1[1] - point2[1]) ** 2)

def filter_elements_within_rect(elements, rect):
    """筛选出所有位于指定矩形内的元素"""
    filtered_elements = [elem for elem in elements if is_within_rect(elem, rect)]
    return filtered_elements

def merge_close_keys(data, threshold):
    # 对键进行排序
    sorted_keys = sorted(data.keys())
    groups = defaultdict(list)
    group_key = sorted_keys[0]

    # 分组相近的键
    for key in sorted_keys:
        if key - group_key <= threshold:
            groups[group_key].append(data[key])
        else:
            group_key = key
            groups[group_key].append(data[key])

    # 合并每个组的值
    merged_data = {k: sum(v) for k, v in groups.items()}
    return merged_data

def init_rows(hor_list):
    y = {}
    for h in hor_list:
        key = round(mid_y(h))
        if key in y.keys():
            y[key].append(h)
        else :
            y[key] = [h]
    table_x1 = y[max(y.keys())][0].x1
    table_x0 = y[max(y.keys())][0].x0
    return y,round(table_x0),round(table_x1)

def init_columns(ver_list):
    x = {}
    for v in ver_list:
        key = round(mid_x(v))
        if key in x.keys():
            x[key].append(v)
        else :
            x[key] = [v]
    table_y1 = x[max(x.keys())][0].y1
    table_y0 = x[max(x.keys())][0].y0
    return x,round(table_y0),round(table_y1)

def line_rank(lines,line):
    for rank,value in enumerate(lines):
        if abs(line-value)<5:
            return rank

def rect_intersects(rect1,rect2):
    """ 检查两个矩形是否相交，矩形由其对角坐标给出 """
    return not (rect1.x1+2 <= rect2.x0 or rect1.x0 >= rect2.x1+2 or
                rect1.y1+2 <= rect2.y0 or rect1.y0 >= rect2.y1+2)

def merge_y_cells(y_keys,rects,key,cell_dict, pdf_path):
    # 同一行，不同列合并
    index_list = [i for i in range(len(y_keys)-1)]
    for r in rects:
        index_list = [item for item in index_list if item not in list(range(line_rank(y_keys,r.y1),line_rank(y_keys,r.y0)))]
    for i in index_list:
        # cell_dict[(key,i)].merge(cell_dict[(key+1,i)])
        try:
            cell_dict[(key,i)].merge(cell_dict[(key+1,i)])
        except KeyError as e:
            logger.error(f"key error for {pdf_path}", e)
            raise Exception(e)

def merge_x_cells(x_keys,rects,key,cell_dict, pdf_path):
    # 同一列，不同行合并
    index_list = [i for i in range(len(x_keys)-1)]
    for r in rects:
        index_list = [item for item in index_list if item not in list(range(line_rank(x_keys,r.x0),line_rank(x_keys,r.x1)))]
    for i in index_list:
        try:
            cell_dict[(i,key)].merge(cell_dict[(i,key+1)])
        except KeyError as e:
            logger.error(f"key error for {pdf_path}", e)
            raise Exception(e)

def get_char_element(element,text=[]):
    if isinstance(element, LTChar):
        # 这里你可以获取每个字符的具体信息，如文本内容、位置等
        # print(f"Character: {element.get_text()}, Font: {element.fontname}, Size: {element.size}, Coordinates: ({element.x0}, {element.y0})")
        text.append(element)
    # 如果元素可以继续分解，继续递归遍历
    if hasattr(element, '_objs'):
        for sub_element in element._objs:
            get_char_element(sub_element,text)

def init_table(ws,rows,columns,table_x0,table_y0,table_x1,table_y1,table_char_list, pdf_path):
    x_keys = sorted(columns.keys()) # 从左往右,垂直线
    column_dict = create_index_to_column_dict(len(x_keys))
    y_keys = sorted(rows.keys(),reverse=True) # 从上往下
    cell_dict = {}
    for x_index,x_key in enumerate(x_keys[:-1]):
        for y_index,y_key in enumerate(y_keys[:-1]):
            cell_dict[(x_index,y_index)] = excel_cell(x_key,y_keys[y_index+1],x_keys[x_index+1],y_key,x_index,y_index)
    for index,key in enumerate(x_keys[1:]):
        rects = columns[key] # 垂直线
        if len(rects)>1 :
            merge_y_cells(y_keys,rects,index,cell_dict, pdf_path)
        elif len(rects) == 0:
            merge_y_cells(y_keys,rects, index,cell_dict, pdf_path)
        elif abs(rects[0].height - (table_y1-table_y0)) > 2:
            merge_y_cells(y_keys,rects,index,cell_dict, pdf_path)
        else:
            pass
    for index,key in enumerate(y_keys[1:]):
        rects = rows[key] # 水平线
        if len(rects)>1 :
            merge_x_cells(x_keys,rects,index,cell_dict, pdf_path)
        elif abs(rects[0].width - (table_x1-table_x0)) > 2:
            merge_x_cells(x_keys,rects,index,cell_dict, pdf_path)
        else:
            pass
    for key in cell_dict.keys():
        cell = cell_dict[key]
        rect = LTRect(bbox=(cell.x0-4,cell.y0-4,cell.x1+4,cell.y1+4),linewidth=1)
        text_list = filter_elements_within_rect(table_char_list,rect)
        text_list = [item.get_text() for item in text_list]
        text = "".join(text_list)
        ws[f"{column_dict[key[0]]}{key[1]+1}"] = text
    for key in cell_dict.keys():
        cell = cell_dict[key]
        if cell.child_cell != []:
            ch =  cell.child_cell[-1]
            ws.merge_cells(f"{column_dict[key[0]]}{key[1]+1}:{column_dict[(ch.column_index)]}{ch.row_index+1}")
            ws[f"{column_dict[key[0]]}{key[1]+1}"].alignment = Alignment(horizontal='center', vertical='center')
    return

def is_within_rect(element, rect):
    # 获取元素和矩形的边界
    ex0, ey0, ex1, ey1 = element.bbox
    rx0, ry0, rx1, ry1 = rect.bbox
    
    # 判断元素是否完全位于矩形内
    return (ex0 >= rx0 and ex1 <= rx1 and ey0 >= ry0 and ey1 <= ry1)

def form_rectangles(points,edge_broaden = 4):
    if len(points) % 4 != 0:
        raise ValueError("The number of points must be a multiple of 4")

    rectangles = []
    for i in range(0, len(points), 4):
        rect_points = points[i:i+4] # rect_point的[0]是右上,[3]是左下
        rectangles.append(LTRect(
                    bbox=(rect_points[3][0]-edge_broaden, rect_points[3][1]-edge_broaden,rect_points[0][0]+edge_broaden,rect_points[0][1]+edge_broaden),linewidth=1))
    return rectangles

def delete_redundant_points(points,threshold = 3):
    for i in range(len(points)):
        for j in range(i + 1, len(points)):
            dist = distance(points[i], points[j])
            if dist <= threshold:
                points[j] = points[i]
    points = list(set(points))
    points = sorted(points, key=lambda point: (point[1], point[0]), reverse=True) # 按从上到下，从右往左排序
    return points

def get_max_area(rect_list,edge_broaden=4):
    min_x0 = 10000
    min_y0 = 10000
    max_x1 = 0
    max_y1 = 0
    for r in rect_list:
        if r.x0<min_x0:
            min_x0 = r.x0
        if r.y0<min_y0:
            min_y0 = r.y0
        if r.x1>max_x1:
            max_x1 = r.x1
        if r.y1>max_y1:
            max_y1 = r.y1
    return LTRect(
                    bbox=(min_x0-edge_broaden,min_y0-edge_broaden,max_x1+edge_broaden,max_y1+edge_broaden),linewidth=1)

def find_table_rect(vertical_list,horizontal_list,threshold):
    rectangles = []
    inter_list = []
    lines_list = vertical_list+horizontal_list
    v_num = len(vertical_list)
    h_num = len(horizontal_list)
    for v_index,v in enumerate(vertical_list):
        for h_index,h in enumerate(horizontal_list):
            if rect_intersects(v,h):
                inter_list.append((v_index,h_index+v_num))
    num_objects = v_num + h_num
    groups = group_intersections(inter_list, num_objects)
    for g in groups:
        if len(g)>=2:
            lines = [lines_list[i] for i in g]
            ver_lines,hor_lines = get_box_line(get_max_area(lines,0))
            vertical_list+=ver_lines
            horizontal_list+= hor_lines
            rectangles.append(get_max_area(lines,threshold))            
    return rectangles  # 如果没有找到任何相近的角

def mid_y(obj):
    x0, y0, x1, y1 = obj.bbox
    mid =  (y0 + y1) / 2
    return mid

def mid_x(obj):
    x0, y0, x1, y1 = obj.bbox
    mid =  (x0 + x1) / 2
    return mid

def get_title_mode(title):
    """Evaluate title format and return a category."""
    # Patterns for identifying different title formats
    patterns = [
        (r"^\s*第[一二三四五六七八九十]{1,}章", 4),
        (r"^\s*第[一二三四五六七八九十]{1,}节", 3),
        (r"^\s*[(（]?[一二三四五六七八九十]{1,}[)）]", 2),
        (r"^\s*[一二三四五六七八九十]{1,}\s*[、\.]", 2),
        (r"^\s*第[1234567890]{1,}章", 4),
        (r"^\s*第[1234567890]{1,}节", 3),
        (r"^\s*[(（]?[1234567890]{1,}[)）]", 5),
        (r"^\s*[1234567890]{1,}\s*[、\.]", 6),
        (r"^\s*(发行概况|发行人声明|重大事项提示|目\s*录|释\s*义|引\s*言)", 3)
    ]
    for pattern, mode in patterns:
        if re.search(pattern, title):
            return mode
    return None

def extract_table_title(char_list, rect, limit=None):
    """
    从指定的表格矩形上方2-65单位内提取文本作为表格名称。
    """
    if limit is None:
        title_region = LTRect(
                        bbox=(rect.x0, rect.y1 - 4, rect.x1, rect.y1 + 40), linewidth=1)  # 设定标题搜索区域
    else:
        title_region = LTRect(
                        bbox=(rect.x0, rect.y1 - 4, rect.x1, rect.y1 + limit), linewidth=1)  # 设定标题搜索区域
    text_list = filter_elements_within_rect(char_list, title_region)
    text_list = [item.get_text() for item in text_list]
    title_text = "".join(text_list)
    title_text = title_text.strip().replace("适用", "").replace("不适用", "").replace("□", "").replace("√", "").replace("不", "").replace("单位：元 币种：人民币", "")

    # 去除Excel不允许的字符(不能作为sheet标题)
    invalid_chars = r'[\\/*?:"\[\]]|:'  # 包括Excel禁止的字符
    cleaned_title = re.sub(invalid_chars, '_', title_text)  # 清洗包括中文、英文、数字和部分符号以外的所有字符

    # 检查标题格式
    title_mode = get_title_mode(cleaned_title)
    if title_mode is not None:
        # 如果标题符合预设格式，返回清洗后的标题
        return cleaned_title
    else:
        # 如果不符合预设格式，返回整个区域内的原始文本，同样需要清洗
        return cleaned_title
    
def group_rectangles(rectangles, mid_func ,threshold=2):
    # 按垂直中点分组
    groups = {}
    for rect in rectangles:
        my = mid_func(rect)
        placed = False
        for key in list(groups.keys()):
            if abs(my - key) < threshold:
                groups[key].append(rect)
                placed = True
                break
        if not placed:
            groups[my] = [rect]
    rect_list = []
    for key in groups.keys():
        rect_list += connect_rectangles(groups[key],mid_func)
    return rect_list

def connect_rectangles(rectangles,mid_func,distance_threshold=2):
    if mid_func == mid_y:
        # 按x0排序矩形
        rectangles.sort(key=lambda rect: rect.bbox[0])

        # 存储连接后的矩形
        connected_rects = []

        # 初始化第一个矩形
        current_rect = rectangles[0]

        for rect in rectangles[1:]:
            # 检查当前矩形的x1是否足够接近下一个矩形的x0
            if current_rect.bbox[2] >= rect.bbox[0] - distance_threshold:  # 有重叠或紧挨着
                # 扩展当前矩形的x1到下一个矩形的x1
                current_rect = LTRect(
                    bbox=(current_rect.bbox[0], current_rect.bbox[1],
                        max(current_rect.bbox[2], rect.bbox[2]),
                        current_rect.bbox[3]),linewidth=1)
            else:
                # 没有重叠，保存当前矩形，并开始新的连接
                connected_rects.append(current_rect)
                current_rect = rect
        
        # 添加最后一个矩形
        connected_rects.append(current_rect)
    if mid_func == mid_x:
        # 按x0排序矩形
        rectangles.sort(key=lambda rect: rect.bbox[1])

        # 存储连接后的矩形
        connected_rects = []

        # 初始化第一个矩形
        current_rect = rectangles[0]

        for rect in rectangles[1:]:
            # 检查当前矩形的x1是否足够接近下一个矩形的x0
            if current_rect.bbox[3] >= rect.bbox[1] - distance_threshold:  # 有重叠或紧挨着
                # 扩展当前矩形的x1到下一个矩形的x1
                current_rect = LTRect(
                    bbox=(current_rect.bbox[0], current_rect.bbox[1],current_rect.bbox[2],
                        max(current_rect.bbox[3], rect.bbox[3])),linewidth=1)
            else:
                # 没有重叠，保存当前矩形，并开始新的连接
                connected_rects.append(current_rect)
                current_rect = rect
        
        # 添加最后一个矩形
        connected_rects.append(current_rect)
    return connected_rects

def copy_alignment(src_cell, dst_cell):
    """仅复制一个单元格的对齐属性，创建新的Alignment对象避免StyleProxy错误"""
    src_alignment = src_cell.alignment
    new_alignment = Alignment(horizontal=src_alignment.horizontal, 
                              vertical=src_alignment.vertical, 
                              text_rotation=src_alignment.text_rotation, 
                              wrap_text=src_alignment.wrap_text,
                              shrink_to_fit=src_alignment.shrink_to_fit,
                              indent=src_alignment.indent)
    dst_cell.alignment = new_alignment

def merge_sheet(sheet1,sheet2):
    # 计算开始插入新行的位置
    start_row = sheet1.max_row + 1
    # 遍历第二个工作表的每一行和每一列
    for i in range(1, sheet2.max_row + 1):
        for j in range(1, sheet2.max_column + 1):
            src_cell = sheet2.cell(row=i, column=j)
            dst_cell = sheet1.cell(row=start_row + i - 1, column=j)
            # 复制单元格的值和样式
            dst_cell.value = src_cell.value
            copy_alignment(src_cell, dst_cell)
    # 处理合并单元格
    for merge_cell in sheet2.merged_cells.ranges:
        new_merge = openpyxl.worksheet.cell_range.CellRange(
            min_row=merge_cell.min_row + start_row - 1,
            min_col=merge_cell.min_col,
            max_row=merge_cell.max_row + start_row - 1,
            max_col=merge_cell.max_col
        )
        sheet1.merge_cells(str(new_merge))

def get_box_line(box):
    ver_lines = []
    hor_lines = []
    ver_lines.append(LTRect(
                    bbox=(box.x0, box.y0, box.x0, box.y1),linewidth=1))
    ver_lines.append(LTRect(
                    bbox=(box.x1, box.y0, box.x1, box.y1),linewidth=1))
    hor_lines.append(LTRect(
                    bbox=(box.x0, box.y0, box.x1, box.y0),linewidth=1))
    hor_lines.append(LTRect(
                    bbox=(box.x0, box.y1, box.x1, box.y1),linewidth=1))
    return ver_lines,hor_lines

def extract_tables_from_pdf(pdf_path, page_number = None):
    # 打开PDF文件
    with open(pdf_path, 'rb') as file:
        parser = PDFParser(file)
        document = PDFDocument(parser)

        wb = Workbook()

        # 创建资源管理器和参数
        rsrcmgr = PDFResourceManager()
        laparams = LAParams()
        device = PDFPageAggregator(rsrcmgr, laparams=laparams)
        interpreter = PDFPageInterpreter(rsrcmgr, device)
        footer_thresholds = 75
        recent_table = None
        # 读取特定页面
        for page_index, page in enumerate(tqdm(PDFPage.create_pages(document))):
            if page_number is None:
                pass
            else :
                if page_index != page_number-1:
                    continue
            interpreter.process_page(page)
            layout = device.get_result()

            vertical_list = []
            horizontal_list = []

            # 遍历页面中的对象
            for obj in layout:
                if isinstance(obj, LTRect):
                    if obj.width<1 and obj.height>2:
                        vertical_list.append(obj)
                    if obj.height<1 and obj.width>2:
                        horizontal_list.append(obj)
                if isinstance(obj, LTLine):
                    if obj.width<1 and obj.height>2:
                        vertical_list.append(obj)
                    if obj.height<1 and obj.width>2:
                        horizontal_list.append(obj)
            if len(vertical_list+horizontal_list) <3 :
                recent_table = None
                continue
            vertical_list = group_rectangles(vertical_list,mid_x)
            horizontal_list = group_rectangles(horizontal_list,mid_y)
            rectangles = find_table_rect(vertical_list,horizontal_list,4)
            if len(rectangles) == 0 :
                recent_table = None
                continue
            vertical_list = group_rectangles(vertical_list,mid_x)
            horizontal_list = group_rectangles(horizontal_list,mid_y)
            rectangles = sorted(rectangles, key=lambda rect: (-rect.y1, rect.x1))
            # # 初始化绘图
            # fig, ax = plt.subplots()
            # for r in rectangles: # 测试表格范围定位是否准确
            #     x0, y0, x1, y1 = r.bbox  # 示例边界值
            #     # 计算矩形的宽度和高度
            #     width = x1 - x0
            #     height = y1 - y0
            #     rect = Rectangle((x0, y0), width, height, linewidth=1, edgecolor='b', facecolor='b')
            #     ax.add_patch(rect)
            # # 测试表格中线条识别情况
            # for v in vertical_list:
            #     mid = mid_x(v)
            #     line = Line2D([mid, mid], [v.y0, v.y1], linewidth=1, color='red')
            #     ax.add_line(line)
            # for h in horizontal_list:
            #     mid = mid_y(h)
            #     line = Line2D([h.x0, h.x1], [mid, mid], linewidth=1, color='red')
            #     ax.add_line(line)
            # # 设置图的边界
            # # 页面尺寸
            # page_width = page.mediabox[2]  # 页面的宽度
            # page_height = page.mediabox[3]  # 页面的高度
            # ax.set_xlim(0, page_width)
            # ax.set_ylim(0, page_height)
            # ax.set_aspect('equal')
            # plt.show()
            # 字体大小一般为8，9
            # 页眉边距一般为75
            page_height = page.mediabox[3] 
            page_width = page.mediabox[2] 
            char_list = [] 
            for element in layout:
                    if isinstance(element, LTTextContainer):
                        get_char_element(element,char_list)
            footer_rect = LTRect(
                        bbox=(0,0,page_width,footer_thresholds),linewidth=1) 
            footer_char = filter_elements_within_rect(char_list,footer_rect)
            footer_list = [item.get_text() for item in footer_char]
            header_rect = LTRect(
                        bbox=(0,page_height-footer_thresholds,page_width,page_height),linewidth=1) 
            header_char = filter_elements_within_rect(char_list,header_rect)
            header_list = [item.get_text() for item in header_char]
            recent_r = None
            for table_index,r in enumerate(rectangles):
                if table_index == 0:
                    before_rect = LTRect(
                        bbox=(0,r.y1,page_width,page_height),linewidth=1) 
                    before_char = filter_elements_within_rect(char_list,before_rect)
                    before_list = [item.get_text() for item in before_char]
                    before_list = [item for item in before_list if item not in header_list]
                    if before_list != []:
                        recent_table = None
                if recent_r:
                    limit = recent_r.y0 - r.y1 + 8
                    if limit >=40 :
                        limit = None
                else:
                    limit = None
                table_char_list = filter_elements_within_rect(char_list,r)
                table_vers = filter_elements_within_rect(vertical_list,r)
                table_vers = group_rectangles(table_vers,mid_x,2)
                table_hors = filter_elements_within_rect(horizontal_list,r)
                table_hors = group_rectangles(table_hors,mid_y,2)
                if table_hors==[] or table_vers == []:
                    recent_table = None
                    continue
                if table_index == 0 and before_list == [] and recent_table is not None: # 需要与上一个sheet合并。
                    diff = recent_table[-1].x0 - r.x0 # 确保两张表的列能够对齐
                    recent_y_keys = [round(key-diff) for key in recent_table[4].keys()]
                    rows,table_x0,table_x1 = init_rows(table_hors)
                    columns,table_y0,table_y1  = init_columns(table_vers)
                    for key in recent_y_keys:
                        if key not in columns.keys():
                            columns[key] = []
                    ws = wb.create_sheet(title=f"temp")
                    try:
                        init_table(ws,rows,columns,table_x0,table_y0,table_x1,table_y1,table_char_list, pdf_path)
                    except Exception as e:
                        logger.error(f"init_table error at page {page_index} for {pdf_path}", e)  
                    merge_sheet(wb[recent_table[2]],wb["temp"])
                    # print(f"merge_sheet {recent_table[2]}")
                    temp_sheet = wb["temp"]
                    wb.remove(temp_sheet)
                    if table_index != len(rectangles)-1 :  
                        recent_table = None
                else:
                    rows,table_x0,table_x1 = init_rows(table_hors)
                    columns,table_y0,table_y1  = init_columns(table_vers)
                    ws = wb.create_sheet(title=f"Page {page_index+1} Table {table_index+1}")
                    if r.y1 >= page_height-75:
                        title = f"Page_{page_index+1} Table_{table_index+1}"
                    else :
                        title = extract_table_title(char_list, r, limit)
                        if not title:  # 如果标题为空，设置默认标题
                            title = f"Page_{page_index+1} Table_{table_index+1}"
                        else :
                            title =  f"Page_{page_index+1} " + title
                    title = title[:30] # 只支持30个字符以内
                    ws.title = title
                    # print(f"Extracted Table Title for Page {page_index+1} Table {table_index+1}: '{title}'")
                    try:
                        init_table(ws,rows,columns,table_x0,table_y0,table_x1,table_y1,table_char_list, pdf_path)
                    except Exception as e:
                        logger.error(f"init_table error at page {page_index} for {pdf_path}", e)
                    if table_index != len(rectangles)-1 :  
                        recent_table = None
                if table_index == len(rectangles)-1 :  
                    after_rect = LTRect(
                        bbox=(0,0,page_width,r.y0),linewidth=1) 
                    after_char = filter_elements_within_rect(char_list,after_rect)
                    after_list = [item.get_text() for item in after_char]
                    after_list = [item for item in after_list if item not in footer_list]
                    if after_list == [] :
                        if recent_table == None :
                            recent_table = (page_index,table_index,title,rows,columns,r)
                        else:
                            recent_table = recent_table 
                    else:
                        recent_table = None
                recent_r = r
        # 移除默认创建的 "Sheet"
        try:
            default_sheet = wb["Sheet"]
            wb.remove(default_sheet)
            wb.save(f'/data/financial_report_baijiu/公司公告/tables/{os.path.basename(pdf_path).split(".pdf")[0]}.xlsx')
            logger.info(f'Saved file {os.path.basename(pdf_path).split(".pdf")[0]}')
        except Exception as e:
            # if os.path.exists(f'/data/financial_report_baijiu/公司公告/tables/{os.path.basename(pdf_path).split(".pdf")[0]}.xlsx'):
            #     os.remove(f'/data/financial_report_baijiu/公司公告/tables/{os.path.basename(pdf_path).split(".pdf")[0]}.xlsx')
            logger.error(f'Fail to save {os.path.basename(pdf_path).split(".pdf")[0]}', e)

def process_pdf(pdf_path):
    extract_tables_from_pdf(pdf_path)


if __name__ == '__main__':
    directory_path = '/data/financial_report_baijiu/公司公告'

    file_path = '/data/shenyuge/lingyue-data-process/pdflux/2024-04-03：贵州茅台：贵州茅台2023年年度报告.pdf'
    extract_tables_from_pdf(file_path)

    pdf_files = []
    processed = set()
    time_outs = []

    for _,_, table in os.walk('/data/financial_report_baijiu/公司公告/tables'):
        for t in table:
            processed.add(t[:-5])

    with open('/data/shenyuge/lingyue-data-process/pdflux/log2024-04-25 12:19.log', 'r') as f:
        while line:= f.readline():
            re_expresssion = re.search(r"Saved file (.+)",line)
            if re_expresssion:
                time_outs += [re_expresssion.groups()[0]]

    for file in glob.glob(f'{directory_path}/**/*.pdf', recursive=True):
        # Check if the file has a PDF extension
        # if (file.endswith('年度报告.pdf') and file.split('/')[-1][:-4] not in processed) or (file in time_outs):
        if file.endswith('年度报告.pdf') and file.split('/')[-1][:-4] not in time_outs:
            # Construct the full path to the PDF file
            pdf_files.append(file)
            # process_pdf(file)

    num_threads = 8  # You can adjust this based on your system's capabilities

    # # Process PDF files using multi-threading
    with ProcessPoolExecutor(max_workers=num_threads) as executor:
        list(tqdm(executor.map(process_pdf, pdf_files), total=len(pdf_files)))
        # {executor.submit(process_pdf, pdf_file): pdf_file for pdf_file in pdf_files}