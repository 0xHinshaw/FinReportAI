import pandas as pd
import openpyxl
import difflib
import re
import logging
from tqdm import tqdm
import os
import json

# 配置日志记录器,修改日志路径可在filename中修改。
logging.basicConfig(filename='/home/luzhenye/PythonProject/表格提取/net_profit.log', level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def get_closest_match(targets, items, cutoff=0.2):
    """获取列表中与目标字符串最相近的项

    Args:
        targets (_type_): 目标字符串
        items (_type_): 检索的列表
        cutoff (float, optional): 相似度阈值，必须高于该阈值才会被考虑. Defaults to 0.2.

    Returns:
        _type_: 返回最相似的匹配项，和匹配项的索引值。没有匹配到则返回None。
    """
    # 初始化最佳匹配和最高相似度
    best_match = None
    highest_score = cutoff  # 只考虑高于cutoff的匹配
    modify_items = [item for item in items if isinstance(item,str)]
    # 遍历每一个目标字符串
    for target in targets:
        # 获取与当前目标字符串相似度最高的匹配项
        matches = difflib.get_close_matches(target, modify_items, n=1, cutoff=cutoff)
        if matches:
            # 检查这个匹配的相似度得分（默认为匹配列表的第一个元素的相似度）
            score = difflib.SequenceMatcher(None, target, matches[0]).ratio()
            # 如果这个得分是目前为止最高的，更新最佳匹配和最高得分
            if score > highest_score:
                best_match = matches
                highest_score = score

    # 如果找到了最佳匹配，打印并返回
    if best_match:
        # print("Found closest match:", best_match)
        return best_match, items.index(best_match[0])
    else:
        # print("No close matches found.")
        return None, -1

def get_table_value(df,row_targets,column_targets):
    """获取表格对应目标行列的数据

    Args:
        df (_type_): dataframe数据
        row_targets (_type_): 目标行
        column_targets (_type_): 目标列

    Returns:
        _type_: 目标行列数据和目标行列数的索引位置。
    """
    close_row,row_index= get_closest_match(row_targets,list(df.iloc[:,0]),0.3)
    close_column,column_index = get_closest_match(column_targets,list(df.columns),0.3)
    value = df.iloc[row_index,column_index]
    return value,(row_index,column_index)

def str2float(string):
    """将表格的数值数据转化成float用于计算。

    Args:
        string (_type_): 传入的数据字符

    Returns:
        _type_: float形式的数据
    """
    return float(re.sub('[\$,]', '',string))

def extra_year_on_year_data(excel_path):
    """获取年度净利润数据和同比信息

    Args:
        excel_path (_type_): excel表路径

    Returns:
        _type_: 输出数据字典
    """
    # 打开Excel文件
    year_pattren = r"([0123456789]{4})(\s*年)"
    year_match = re.search(year_pattren,excel_path)
    if year_match:
        recent_year = year_match.group()
        previous_year = "".join(str(int(year_match.group(1))-1)+"年")
    else:
        raise "年份识别失败"
    target_table = ["合并利润"]
    wb = openpyxl.load_workbook(excel_path)
    # 获取所有工作表的名称
    sheet_names = wb.sheetnames
    close_sheet,sheet_index = get_closest_match(target_table,sheet_names)
    if isinstance(close_sheet,list):
        close_sheet = close_sheet[0]
    if close_sheet != None: #检查匹配到的表格是否正确
        df = pd.read_excel(excel_path,close_sheet)
        t_row,t_row_index = get_closest_match(["归属于母公司股东的净利润"],list(df.iloc[:,0]),0.4)
        if t_row :
            t_column,t_column_index = get_closest_match(["本期发生额","当期",f"{recent_year}"],list(df.columns),0.2)
            if t_column:
                pass
            else :close_sheet = None
        else:close_sheet = None
    if close_sheet == None: # 如果没匹配到对应表格名字，则对所有表格的行列进行匹配。
        for t in sheet_names:
            df = pd.read_excel(excel_path,t)
            if df.empty:
                continue
            t_row,t_index = get_closest_match(["归属于母公司股东的净利润"],list(df.iloc[:,0]),0.5)
            if t_row :
               if t_row :
                t_column,t_column_index = get_closest_match(["本期发生额","当期",f"{recent_year}"],list(df.columns),0.2)
                if t_column:
                    close_sheet = t
                    break
    if close_sheet == None:
        logging.info(f"找不到 归属于母公司股东的净利润 季度数据 相关内容:{excel_path}")
        return
    if isinstance(close_sheet,list):
        close_sheet = close_sheet[0]
    sheet_name=f'{str(close_sheet)}'
    # print(sheet_name)
    # 指定工作表名称
    df = pd.read_excel(excel_path, sheet_name)
    curent_value,curent_value_index = get_table_value(df,["归属于母公司股东的净利润"],["本期发生额","当期",f"{recent_year}"])
    previous_value,curent_value_index = get_table_value(df,["归属于母公司股东的净利润"],["上期发生额","上期",f"{previous_year}"])
    # print(df.iloc[curent_value_index[0]:curent_value_index[0]+1,:])
    if str(previous_value) != "nan" and str(curent_value) != "nan":
        return {"本期：":curent_value,"上期：":previous_value,"同比：":round(str2float(curent_value)/str2float(previous_value) -1, 4) }
    else:
        return {"本期：":curent_value,"上期：":previous_value,"同比：":"nan" }

def extra_recent_data(excel_path):
    """获取当期的总量数据

    Args:
        excel_path (_type_): excel的表格路径

    Returns:
        _type_: 当期数据
    """
    # 打开Excel文件
    year_pattren = r"([0123456789]{4})(\s*年)"
    year_match = re.search(year_pattren,excel_path)
    if year_match:
        recent_year = year_match.group()
        previous_year = "".join(str(int(year_match.group(1))-1)+"年")
    else:
        raise "年份识别失败"
    target_table = ["合并利润"]
    wb = openpyxl.load_workbook(excel_path)
    # 获取所有工作表的名称
    sheet_names = wb.sheetnames
    close_sheet,sheet_index = get_closest_match(target_table,sheet_names)
    if isinstance(close_sheet,list):
        close_sheet = close_sheet[0]
    if close_sheet != None: # 检查匹配到的表格是否正确
        df = pd.read_excel(excel_path,close_sheet)
        t_row,t_row_index = get_closest_match(["归属于母公司股东的净利润"],list(df.iloc[:,0]),0.4)
        if t_row :
            t_column,t_column_index = get_closest_match(["本期发生额","当期",f"{recent_year}"],list(df.columns),0.2)
            if t_column:
                pass
            else :close_sheet = None
        else:close_sheet = None
    if close_sheet == None: # 如果没匹配到对应表格名字，则对所有表格的行列进行匹配。
        for t in sheet_names:
            df = pd.read_excel(excel_path,t)
            if df.empty:
                continue
            t_row,t_row_index = get_closest_match(["归属于母公司股东的净利润"],list(df.iloc[:,0]),0.5)
            if t_row :
               if t_row :
                t_column,t_column_index = get_closest_match(["本期发生额","当期",f"{recent_year}"],list(df.columns),0.2)
                if t_column:
                    close_sheet = t
                    break
    if close_sheet == None:
        logging.info(f"找不到 归属于母公司股东的净利润 当期数据 相关内容{excel_path}")
        return
    if isinstance(close_sheet,list):
        close_sheet = close_sheet[0]
    sheet_name=f'{str(close_sheet)}'
    # print(sheet_name)
    # 指定工作表名称
    df = pd.read_excel(excel_path, sheet_name)
    curent_value,curent_value_index = get_table_value(df,["归属于母公司股东的净利润"],["本期发生额","当期",f"{recent_year}"])
    # print(df.iloc[curent_value_index[0]:curent_value_index[0]+1,:])
    return {"本期：":curent_value }

def extra_quarter_over_quarter_data(excel_path):
    """获取季度环比数据和相应的季度数据

    Args:
        excel_path (_type_): excel表路径

    Returns:
        _type_: 字典包含各个季度的值以及后三个季度的环比。
    """
    year_pattren = r"([0123456789]{4})(\s*年)"
    year_match = re.search(year_pattren,excel_path)
    if year_match:
        recent_year = year_match.group()
        previous_year = "".join(str(int(year_match.group(1))-1)+"年")
    else:
        raise "年份识别失败"
    if re.search("半年",excel_path):
        target_table = ["主要会计数据"]
        columns_list = [["本报告期",f"{recent_year}"],["上年同期",f"{previous_year}"]]
    else:
        target_table = ["分季度主要财务数据"]
        columns_list = [["第一季度"],["第二季度"],["第三季度"],["第四季度"]]

    # 打开Excel文件
    wb = openpyxl.load_workbook(excel_path)
    # 获取所有工作表的名称
    sheet_names = wb.sheetnames
    close_sheet,sheet_index = get_closest_match(target_table,sheet_names)
    if isinstance(close_sheet,list):
        close_sheet = close_sheet[0]
    if close_sheet != None: #检查匹配到的表格是否正确
        df = pd.read_excel(excel_path,close_sheet)
        t_row,t_row_index = get_closest_match(["归属于母公司股东的净利润"],list(df.iloc[:,0]),0.4)
        if t_row :
            t_column,t_column_index = get_closest_match(columns_list[0],list(df.columns),0.2)
            if t_column:
                pass
            else :close_sheet = None
        else:close_sheet = None
    if close_sheet == None:# 如果没匹配到对应表格名字，则对所有表格的行列进行匹配。
        for t in sheet_names:
            df = pd.read_excel(excel_path,t)
            if df.empty:
                continue
            t_row,t_row_index = get_closest_match(["归属于母公司股东的净利润"],list(df.iloc[:,0]),0.4)
            if t_row :
                t_column,t_column_index = get_closest_match(columns_list[0],list(df.columns),0.2)
                if t_column:
                    close_sheet = t
                    break
    if close_sheet == None:
        logging.info(f"找不到 归属于母公司股东的净利润 季度数据 相关内容:{excel_path}")
        return
    if isinstance(close_sheet,list):
        close_sheet = close_sheet[0]
    sheet_name=f'{str(close_sheet)}'
    # print(sheet_name)
    # 指定工作表名称
    df = pd.read_excel(excel_path, sheet_name)
    quarter_data = []
    for c in columns_list:
        quarter_value,quarter_value_index = get_table_value(df,["归属于母公司股东的净利润"],c)
        if quarter_value:
            quarter_data.append(quarter_value)
    # print(df.iloc[quarter_value_index[0]:quarter_value_index[0]+1,:])
    growth_rates = []
    for i in range(1, len(quarter_data)):
        if str(quarter_data[i]) != "nan" and str(quarter_data[i-1]) != "nan":
            growth_rate = (str2float(quarter_data[i]) - str2float(quarter_data[i - 1])) / str2float(quarter_data[i - 1]) 
            growth_rates.append(round(growth_rate,4))
        else :  growth_rates.append("nan")
    # 输出每个季度的环比增长率
    if re.search("半年",excel_path):
        if str(quarter_data[0]) != "nan" and str(quarter_data[1]) != "nan":
            return {"本期：":quarter_data[0],"上期：":quarter_data[1],"同比：":round(str2float(quarter_data[0])/str2float(quarter_data[1]) -1, 4) }
        else:
            return {"本期：":quarter_data[0],"上期：":quarter_data[1],"同比：":"nan"}
    else:
        # for i, rate in enumerate(growth_rates, start=1):
        #     print(f"从第{i}季度到第{i+1}季度的环比增长率是: {rate:.4f}")
        return {"第一季度：":quarter_data[0],"第二季度：":quarter_data[1],"第三季度":quarter_data[2],"第四季度":quarter_data[3],"季度环比":growth_rates}

def get_net_profit(excel_path,target):
    """假设用户只传入他所需要的指标名称，通过字典对应的函数，调用函数返回结果。

    Args:
        excel_path (_type_): excel表路径
        target (str): 用户要求的指标，对应target_dict里面的函数

    Returns:
        _type_: 返回一个字典，包含指标的内容
    """
    """假设用户只传入他所需要的指标名称，通过字典对应的函数，调用函数返回结果。

    Args:
        excel_path (_type_): excel表路径
        target (str): 用户要求的指标，对应target_dict里面的函数

    Returns:
        _type_: 返回一个字典，包含指标的内容
    """
    target_dict = {
            "同比":extra_year_on_year_data,
            "总量":extra_recent_data,
            "季度环比":extra_quarter_over_quarter_data
        }
    func = target_dict[target]
    return func(excel_path)

if __name__ == "__main__" :
    # excels_path = os.listdir("/data/financial_report_baijiu/公司公告/tables")
    # output_dir = "/home/luzhenye/PythonProject/表格提取/net_profit_data"
    # for excel_path in tqdm(excels_path) :
    #     file_name = os.path.splitext(os.path.basename(excel_path))[0] #文件名字
    #     json_file_path = os.path.join(output_dir,file_name+".json")  
    #     result = {} 
    #     try: # 创建输出目录
    #         excel_path = os.path.join("/data/financial_report_baijiu/公司公告/tables",excel_path)
    #         result["同比"] = get_net_profit(excel_path,"同比")
    #         result["总量"] = get_net_profit(excel_path,"总量")
    #         result["季度环比"] = get_net_profit(excel_path,"季度环比")
    #         with open(json_file_path, 'w', encoding='utf-8') as f:
    #             json.dump(result, f, ensure_ascii=False, indent=4)
    #     except Exception as e:
    #         logging.info("转换失败:{}".format(excel_path)) # 转换失败记录
    excel_path = "2019-08-30：＊ST皇台：2019年半年度报告.xlsx"
    file_name = os.path.splitext(os.path.basename(excel_path))[0] #文件名字
    output_dir = "/home/luzhenye/PythonProject/表格提取/net_profit_data"
    json_file_path = os.path.join(output_dir,file_name+".json")  
    result = {}
    excel_path = os.path.join("/data/financial_report_baijiu/公司公告/tables",excel_path)
    result["同比"] = get_net_profit(excel_path,"同比")
    result["总量"] = get_net_profit(excel_path,"总量")
    result["季度环比"] = get_net_profit(excel_path,"季度环比")
    with open(json_file_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=4)

        