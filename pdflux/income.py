import openpyxl
import pandas as pd
import re
import math
import os

# Function to analyze annual income growth
def analyze_income(file_path, current_year, report_type):
    file_name = os.path.basename(file_path)
    wb = openpyxl.load_workbook(file_path)
    target_sheets = ["合并利润表", "主要会计数据", "主要财务数据"]
    ws = None

    for sheet_name in wb.sheetnames:
        if "合并利润表" in sheet_name:
            ws = wb[sheet_name]
            break

    if not ws:
        for sheet_name in wb.sheetnames:
            if any(kw in sheet_name for kw in target_sheets[1:]):
                ws = wb[sheet_name]
                break

    if not ws:
        raise ValueError("没有找到包含'合并利润表', '主要会计数据', 或 '主要财务数据'的sheet名称")

    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=10, values_only=True):
        if row and any(cell and (f"{current_year}年{report_type}" in str(cell) or f"{current_year}{report_type}" in str(cell) or f"{current_year}年" in str(cell) or "本报告期" in str(cell) or "本期发生额" in str(cell) or "上期发生额" in str(cell) or "上年同期" in str(cell)) for cell in row):
            header_row = row
            break

    if header_row is None:
        raise ValueError("无法找到合适的表头行")

    income_row = find_income_row(ws)
    if income_row is None:
        return {"error": "未找到'营业总收入'或'营业收入'的行"}

    current_year_col = next((i for i, cell in enumerate(header_row) if any(label in str(cell) for label in [f"{current_year}年{report_type}",f"{current_year}年",f"{current_year}{report_type}","本期发生额", "本报告期"])), None)
    previous_year_col = next((i for i, cell in enumerate(header_row) if any(label in str(cell) for label in [f"{current_year - 1}年{report_type}",f"{current_year}年", f"{current_year}{report_type}","上期发生额", "上年同期"])), None)

    if current_year_col is None or previous_year_col is None:
        raise ValueError("财务数据列索引错误")

    current_year_income = float(income_row[current_year_col].replace(',', ''))
    previous_year_income = float(income_row[previous_year_col].replace(',', ''))

    def calculate_growth(current, previous):
        return (current - previous) / previous if previous != 0 else float('nan')

    growth_rate = calculate_growth(current_year_income, previous_year_income)
    return {"营业总收入": current_year_income, "同比增长率": growth_rate}

def find_income_row(worksheet):
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            if "营业总收入" in row[0]:
                return row
            elif "营业收入" in row[0]:
                return row
    return None

# Function to calculate quarterly growth
def calculate_quarterly_growth(file_path):
    xls = pd.ExcelFile(file_path)
    quarterly_sheets = [sheet_name for sheet_name in xls.sheet_names if '季度' in sheet_name]

    results = {}  # To store quarterly growth results

    for sheet_name in quarterly_sheets:
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        df.columns = ['项目', '一季度', '二季度', '三季度', '四季度']

        for column in df.columns[1:]:
            df[column] = df[column].apply(lambda x: float(re.sub(r'[^\d.]', '', str(x))) if isinstance(x, str) else x)

        for i in range(2, 5):  # Calculating from the second quarter
            quarter_name = df.columns[i]
            last_quarter = df.columns[i - 1]
            quarter_growth = ((df[quarter_name] - df[last_quarter]) / df[last_quarter]) * 100
            results[f"{quarter_name} 季环比"] = f"{quarter_growth.iloc[0]:.2f}%"

    return results

def get_income_info(file_path, parameter):
    assert parameter in ["营业总收入总量",'营业总收入同比', '营业总收入季环比'], "invalid parameter"
    try:
        year_match = re.search(r'(\d{4})年(半年度|年度)报告', file_path)
        if not year_match:
            print(f"文件：{file_path}，错误：无法从文件名中提取年份和报告类型。")
            return
        current_year = int(year_match.group(1))
        report_type = year_match.group(2)
        income_result = analyze_income(file_path, current_year, report_type)
        quarterly_results = calculate_quarterly_growth(file_path)
        income_result.update(quarterly_results)
        if parameter == "营业总收入总量":
            return {"营业总收入总量": income_result.get('营业总收入', None)}
        elif parameter == '营业总收入同比':
            return {'营业总收入同比': income_result.get('同比增长率', None)}
        else:
            return quarterly_results
    except Exception as e:
        print(f"文件：{file_path}，错误：{str(e)}")


# Main loop to process files
if __name__ == '__main__':
    directory = r"/data/financial_report_baijiu/公司公告/tables"
    for filename in os.listdir(directory):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(directory, filename)
            try:
                year_match = re.search(r'(\d{4})年(半年度|年度)报告', filename)
                if not year_match:
                    print(f"文件：{filename}，错误：无法从文件名中提取年份和报告类型。")
                    continue

                current_year = int(year_match.group(1))
                report_type = year_match.group(2)
                income_result = analyze_income(file_path, current_year, report_type)
                quarterly_results = calculate_quarterly_growth(file_path)
                income_result.update(quarterly_results)
                print(f"文件：{filename}，结果：{income_result}")
            except Exception as e:
                print(f"文件：{filename}，错误：{str(e)}")
