from datetime import datetime
from collections import defaultdict
import logging
from logging import Logger
import signal
import re
from concurrent.futures import ProcessPoolExecutor
import glob
from pdfminer.converter import PDFPageAggregator
from matplotlib.patches import Rectangle
from matplotlib.lines import Line2D
from pdfminer.layout import LAParams, LTTextBox, LTTextLine, LTImage, LTLine, LTRect, LTChar, LTTextBoxHorizontal,LTTextContainer
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.pdfpage import PDFPage
from pdfminer.pdfparser import PDFParser
from pdfminer.pdfdocument import PDFDocument
import matplotlib.pyplot as plt
from openpyxl import Workbook
from math import sqrt
import openpyxl
from cell_class import excel_cell
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import os
from tqdm import tqdm
from collections import defaultdict

logger = Logger(log_file_path=f'/home/luzhenye/PythonProject/表格提取/logs/extract_table_log{datetime.now().strftime("%Y-%m-%d %H:%M")}.log')

class UnionFind:
    """该类用于合并同所属同一个表格的线条。"""
    def __init__(self, size):
        self.parent = list(range(size))
        self.rank = [1] * size

    def find(self, p):
        if self.parent[p] != p:
            self.parent[p] = self.find(self.parent[p])  # 路径压缩
        return self.parent[p]

    def union(self, p, q):
        rootP = self.find(p)
        rootQ = self.find(q)
        if rootP != rootQ:
            if self.rank[rootP] > self.rank[rootQ]:
                self.parent[rootQ] = rootP
            elif self.rank[rootP] < self.rank[rootQ]:
                self.parent[rootP] = rootQ
            else:
                self.parent[rootQ] = rootP
                self.rank[rootP] += 1

    def connected(self, p, q):
        return self.find(p) == self.find(q)
   
def group_intersections(intersections, num_objects):
    """有相交关系的线条均划为一组

    Args:
        intersections (_type_): 相交关系，单个相交关系形式为：（1，2），表示线条1与线条2相交。
        num_objects (_type_): 总共有多少线条。

    Returns:
        _type_: 返回一个列表，列表中每一组线条被存放在列表中：例如：[[1,2],[3,4,5,6],[7,8]]
    """
    uf = UnionFind(num_objects)  # 创建并查集，大小为 num_objects，假设最大编号为 num_objects-1
    for i, j in intersections:
        uf.union(i, j)
    
    groups = {}
    for object_id in range(num_objects):
        root = uf.find(object_id)
        if root not in groups:
            groups[root] = []
        groups[root].append(object_id)
    
    return list(groups.values())

def modify_bbox(bbox):
    """用于调整一个rect对象的位置坐标，将其四舍五入为整数。

    Args:
        bbox (_type_): rect对象的bbox属性

    Returns:
        _type_: 返回四舍五入后的bbox属性
    """
    for index,item in enumerate(bbox):
        bbox[index] = round(item)
    return bbox

def create_index_to_column_dict(n):
    """
    创建一个从整数索引到Excel列字母的映射字典。
    参数:
        n (int): 最大索引数，例如，如果n=28，将生成从1到28的映射。
    返回:
        dict: 映射字典，键为整数，值为对应的Excel列字母。
    """
    return {i-1: get_column_letter(i) for i in range(1, n + 1)}

def distance(point1, point2):
    """计算两点距离

    Args:
        point1 (_type_): 点1
        point2 (_type_): 点2

    Returns:
        _type_: 点距
    """
    """计算两点距离

    Args:
        point1 (_type_): 点1
        point2 (_type_): 点2

    Returns:
        _type_: 点距
    """
    return sqrt((point1[0] - point2[0]) ** 2 + (point1[1] - point2[1]) ** 2)

def filter_elements_within_rect(elements, rect):
    """筛选出所有位于指定矩形内的元素"""
    filtered_elements = [elem for elem in elements if is_within_rect(elem, rect)]
    return filtered_elements

def merge_close_keys(data, threshold):
    """合并字典中key值相近的数据

    Args:
        data (_type_): 字典
        threshold (_type_): 判断是否相近的阈值

    Returns:
        _type_: 合并后的字典。
    """
    """合并字典中key值相近的数据

    Args:
        data (_type_): 字典
        threshold (_type_): 判断是否相近的阈值

    Returns:
        _type_: 合并后的字典。
    """
    # 对键进行排序
    sorted_keys = sorted(data.keys())
    groups = defaultdict(list)
    group_key = sorted_keys[0]

    # 分组相近的键
    for key in sorted_keys:
        if key - group_key <= threshold:
            groups[group_key].append(data[key])
        else:
            group_key = key
            groups[group_key].append(data[key])

    # 合并每个组的值
    merged_data = {k: sum(v) for k, v in groups.items()}
    return merged_data

def init_rows(hor_list):
    """根据水平线列表，创建一个字典，字典的键是纵坐标，值是处于同一纵坐标的水平线。

    Args:
        hor_list (_type_): 水平线列表

    Returns:
        _type_: 纵坐标对应水平线字典，以及所有水平线的横坐标范围，表示表格的横坐标范围。
    """
    """根据水平线列表，创建一个字典，字典的键是纵坐标，值是处于同一纵坐标的水平线。

    Args:
        hor_list (_type_): 水平线列表

    Returns:
        _type_: 纵坐标对应水平线字典，以及所有水平线的横坐标范围，表示表格的横坐标范围。
    """
    y = {}
    for h in hor_list:
        key = round(mid_y(h))
        if key in y.keys():
            y[key].append(h)
        else :
            y[key] = [h]
    table_x1 = y[max(y.keys())][0].x1
    table_x0 = y[max(y.keys())][0].x0
    return y,round(table_x0),round(table_x1)

def init_columns(ver_list):
    """根据垂直线列表，创建一个字典，字典的键是横坐标，值是处于同一横坐标的垂直线。

    Args:
        ver_list (_type_): 垂直线列表

    Returns:
        _type_: 横坐标对应垂直线字典，以及所有垂直线的纵坐标范围，表示表格的纵坐标范围。
    """
    """根据垂直线列表，创建一个字典，字典的键是横坐标，值是处于同一横坐标的垂直线。

    Args:
        ver_list (_type_): 垂直线列表

    Returns:
        _type_: 横坐标对应垂直线字典，以及所有垂直线的纵坐标范围，表示表格的纵坐标范围。
    """
    x = {}
    for v in ver_list:
        key = round(mid_x(v))
        if key in x.keys():
            x[key].append(v)
        else :
            x[key] = [v]
    table_y1 = x[max(x.keys())][0].y1
    table_y0 = x[max(x.keys())][0].y0
    return x,round(table_y0),round(table_y1)

def in_keys(lines,line):
    """用于判断某一点是否在列表中，且误差不大。

    Args:
        lines (_type_): 线列表
        line (_type_): 目标线

    Returns:
        _type_: 目标线在线列表中的排序
    """
    # 计算每个元素与目标值的绝对差异
    differences = [abs(x - line) for x in lines if abs(x - line)<7 ]
    # 找出差异最小的索引
    if differences == []:
        return None
    else :
        return differences

def line_rank(lines,line):
    """用于计算某一线在一个列表中排序。

    Args:
        lines (_type_): 线列表
        line (_type_): 目标线

    Returns:
        _type_: 目标线在线列表中的排序
    """
    for rank,value in enumerate(lines):
        if abs(line-value)<5:
            return rank

def rect_intersects(rect1,rect2):
    """ 检查两个矩形是否相交，矩形由其对角坐标给出 """
    return not (rect1.x1+2 <= rect2.x0 or rect1.x0 >= rect2.x1+2 or
                rect1.y1+2 <= rect2.y0 or rect1.y0 >= rect2.y1+2)

def merge_y_cells(y_keys,rects,key,cell_dict):
    """根据线条信息，合并单元格

    Args:
        y_keys (_type_): 纵坐标列表
        rects (_type_): 垂直线
        key (_type_): 列索引
        cell_dict (_type_): 单元格字典
    """
    # 同一行，不同列合并
    index_list = [i for i in range(len(y_keys)-1)]
  
    for r in rects:
        index_list = [item for item in index_list if item not in list(range(line_rank(y_keys,r.y1),line_rank(y_keys,r.y0)))]
    for i in index_list:
        cell_dict[(i,key)].merge(cell_dict[(i,key+1)])

def merge_x_cells(x_keys,rects,key,cell_dict):
    """根据线条信息，合并单元格

    Args:
        x_keys (_type_): 横坐标列表
        rects (_type_): 水平线
        key (_type_): 行索引
        cell_dict (_type_): 单元格字典
    """
    # 同一列，不同行合并
    index_list = [i for i in range(len(x_keys)-1)]
    for r in rects:
        index_list = [item for item in index_list if item not in list(range(line_rank(x_keys,r.x0),line_rank(x_keys,r.x1)))]
    for i in index_list:
        cell_dict[(key,i)].merge(cell_dict[(key+1,i)])

def get_char_element(element,text):
    """遍历element，获取其中的TLChar元素。

    Args:
        element (_type_): 用于遍历检索的元素
        text (list): 用于存放TLChar的列表 .
    """
    if isinstance(element, LTChar):
        # 这里你可以获取每个字符的具体信息，如文本内容、位置等
        # print(f"Character: {element.get_text()}, Font: {element.fontname}, Size: {element.size}, Coordinates: ({element.x0}, {element.y0})")
        text.append(element)
    # 如果元素可以继续分解，继续递归遍历
    if hasattr(element, '_objs'):
        for sub_element in element._objs:
            get_char_element(sub_element,text)

def init_table(ws,rows,columns,table_x0,table_y0,table_x1,table_y1,table_char_list):
    x_keys = sorted(columns.keys()) # 从左往右,垂直线
    column_dict = create_index_to_column_dict(len(x_keys))
    y_keys = sorted(rows.keys(),reverse=True) # 从上往下，纵坐标列表
    # 清洗垂直线和水平线
    for index,key in enumerate(x_keys[1:]):
        rm_list = []
        for r in columns[key]: # 垂直线
            if line_rank(y_keys,r.y1) is None or line_rank(y_keys,r.y0) is None:
                rm_list.append(r)
        columns[key] = [item for item in columns[key] if item not in rm_list]
        if columns[key] == [] :
            del columns[key]
    for index,key in enumerate(y_keys[1:]):
        rm_list = []
        for r in rows[key]: # 水平线
            if line_rank(x_keys,r.x1) is None or line_rank(x_keys,r.x0) is None:
                rm_list.append(r)
        rows[key] = [item for item in rows[key] if item not in rm_list]
        if rows[key] == [] :
            del rows[key]
    cell_dict = {}
    # 更新清洗完的结果
    x_keys = sorted(columns.keys())
    y_keys = sorted(rows.keys(),reverse=True) 
    for x_index,x_key in enumerate(x_keys[:-1]):
        for y_index,y_key in enumerate(y_keys[:-1]):
            cell_dict[(y_index,x_index)] = excel_cell(x_key,y_keys[y_index+1],x_keys[x_index+1],y_key,y_index,x_index)
    for index,key in enumerate(x_keys[1:]):
        rects = columns[key] # 垂直线
        if len(rects)>1 :
            merge_y_cells(y_keys,rects,index,cell_dict)
        elif len(rects) == 0:
            merge_y_cells(y_keys,rects, index,cell_dict)
        elif abs(rects[0].height - (table_y1-table_y0)) > 2:
            merge_y_cells(y_keys,rects,index,cell_dict)
        else:
            pass
    for index,key in enumerate(y_keys[1:]):
        rects = rows[key] # 水平线
        if len(rects)>1 :
            merge_x_cells(x_keys,rects,index,cell_dict)
        elif len(rects) == 0:
            merge_y_cells(y_keys,rects, index,cell_dict)
        elif abs(rects[0].width - (table_x1-table_x0)) > 2:
            merge_x_cells(x_keys,rects,index,cell_dict)
        else:
            pass
    for key in cell_dict.keys():
        cell = cell_dict[key]
        rect = LTRect(bbox=(cell.x0-4,cell.y0-4,cell.x1+4,cell.y1+4),linewidth=1)
        text_list = filter_elements_within_rect(table_char_list,rect)
        text_list = [item.get_text() for item in text_list]
        text = "".join(text_list)
        ws[f"{column_dict[key[1]]}{key[0]+1}"] = text
    for key in cell_dict.keys():
        cell = cell_dict[key]
        if cell.child_cell != []:
            ch =  cell.child_cell[-1]
            ws.merge_cells(f"{column_dict[key[1]]}{key[0]+1}:{column_dict[(ch.column_index)]}{ch.row_index+1}")
            ws[f"{column_dict[key[1]]}{key[0]+1}"].alignment = Alignment(horizontal='center', vertical='center')
    return

def is_within_rect(element, rect):
    """判断element是否在rect范围内

    Args:
        element (_type_): _description_
        rect (_type_): _description_

    Returns:
        _type_: _description_
    """
    # 获取元素和矩形的边界
    ex0, ey0, ex1, ey1 = element.bbox
    rx0, ry0, rx1, ry1 = rect.bbox
    
    # 判断元素是否完全位于矩形内
    return (ex0 >= rx0 and ex1 <= rx1 and ey0 >= ry0 and ey1 <= ry1)

def form_rectangles(points,edge_broaden = 4):
    """根据表格四个角的坐标构建rect，这个方案已经废弃。

    Args:
        points (_type_): _description_
        edge_broaden (int, optional): _description_. Defaults to 4.

    Raises:
        ValueError: _description_

    Returns:
        _type_: _description_
    """
    if len(points) % 4 != 0:
        raise ValueError("The number of points must be a multiple of 4")

    rectangles = []
    for i in range(0, len(points), 4):
        rect_points = points[i:i+4] # rect_point的[0]是右上,[3]是左下
        rectangles.append(LTRect(
                    bbox=(rect_points[3][0]-edge_broaden, rect_points[3][1]-edge_broaden,rect_points[0][0]+edge_broaden,rect_points[0][1]+edge_broaden),linewidth=1))
    return rectangles

def delete_redundant_points(points,threshold = 3):
    """删除冗余的点，这个方案已经废弃。

    Args:
        points (_type_): _description_
        threshold (int, optional): _description_. Defaults to 3.

    Returns:
        _type_: _description_
    """
    for i in range(len(points)):
        for j in range(i + 1, len(points)):
            dist = distance(points[i], points[j])
            if dist <= threshold:
                points[j] = points[i]
    points = list(set(points))
    points = sorted(points, key=lambda point: (point[1], point[0]), reverse=True) # 按从上到下，从右往左排序
    return points

def get_max_area(rect_list,edge_broaden=4):
    """根据某一组线列表，提取出所有线坐标能组成的最大范围的rect。

    Args:
        rect_list (_type_): 某一组线组成的列表
        edge_broaden (int, optional): 边缘拓展. Defaults to 4.

    Returns:
        _type_: 根据最大范围创建的rect对象。
    """
    min_x0 = 10000
    min_y0 = 10000
    max_x1 = 0
    max_y1 = 0
    for r in rect_list:
        if r.x0<min_x0:
            min_x0 = r.x0
        if r.y0<min_y0:
            min_y0 = r.y0
        if r.x1>max_x1:
            max_x1 = r.x1
        if r.y1>max_y1:
            max_y1 = r.y1
    return LTRect(
                    bbox=(min_x0-edge_broaden,min_y0-edge_broaden,max_x1+edge_broaden,max_y1+edge_broaden),linewidth=1)

def find_table_rect(vertical_list,horizontal_list,threshold):
    """获取当前页面所有表格的区域rects

    Args:
        vertical_list (_type_): 垂直线list
        horizontal_list (_type_): 水平线list
        threshold (_type_): 边界拓展阈值

    Returns:
        _type_: rect组成的列表，包含从上到下所有的表格区域。
    """
    rectangles = []
    inter_list = []
    valid_groups = []
    lines_list = vertical_list+horizontal_list
    v_num = len(vertical_list)
    h_num = len(horizontal_list)
    for v_index,v in enumerate(vertical_list):
        for h_index,h in enumerate(horizontal_list):
            if rect_intersects(v,h):
                inter_list.append((v_index,h_index+v_num))
    num_objects = v_num + h_num
    groups = group_intersections(inter_list, num_objects)
    for g in groups:
        if len(g)>=2:
            lines = [lines_list[i] for i in g]
            ver_lines,hor_lines = get_box_line(get_max_area(lines,0))
            vertical_list+=ver_lines
            horizontal_list+= hor_lines
            rectangles.append(get_max_area(lines,threshold))    
        else :
            if g[0] <v_num:
                vertical_list.remove(lines_list[g[0]])
            else:
                horizontal_list.remove(lines_list[g[0]])
    return rectangles  # 如果没有找到任何相近的角

def mid_y(obj):
    """获取某个rect的纵坐标中点

    Args:
        obj (_type_): 某一rect

    Returns:
        _type_: 纵坐标中点
    """
    x0, y0, x1, y1 = obj.bbox
    mid =  (y0 + y1) / 2
    return mid

def mid_x(obj):
    """获取某个rect的横坐标中点

    Args:
        obj (_type_): 某一rect

    Returns:
        _type_: 横坐标中点
    """
    x0, y0, x1, y1 = obj.bbox
    mid =  (x0 + x1) / 2
    return mid

def get_content_title_mode(content, only_mode=True):
    """在检测content内容是否包含标题的函数，返回检测出的第一个标题的格式和检测的结果。
    已经弃用

    Args:
        content (str): 检测内容
        only_mode (bool, optional): 是否只返回标题格式分类. Defaults to True.

    Returns:
        _type_: 第一个匹配成功的标题的格式分类，第一个匹配成功的标题的内容，第一个匹配成功的标题的结束位置索引。
    """
    first_pattern = None
    first_match = None
    first_match_start = 0
    first_match_end = 0
    patterns = [r"[\n\s:：]*[(（]?[一二三四五六七八九十]{1,}[)）]",
                r"[\n\s:：]*[一二三四五六七八九十]{1,}\s*[、\.]",
                r"[\n\s:：]*[(（][l1234567890]{1,2}[)）]",
                r"[\n\s:：]*[l1234567890]{1,2}\s*[、\.](?![1234567890])",
                r"\n\s*项目[l1234567890一二三四五六七八九十]{1,2}\s*[:：]",
                r"\n\s*附送[l1234567890一二三四五六七八九十]{1,2}\s*[:：]",
                r"[\n\s]+(.*?)(表)"]
    for pattern in patterns:
        match = re.search(pattern, content, re.DOTALL)
        if match and match.start() >= first_match_start:
            first_match_start = match.start()
            first_match = match.group()
            first_match_end = match.end()
            first_pattern = patterns.index(pattern)
    if only_mode:
        return first_pattern
    return first_pattern, first_match, first_match_start

def get_title_mode(title):
    """Evaluate title format and return a category."""
    # Patterns for identifying different title formats
    patterns = [
        (r"^\s*第[一二三四五六七八九十]{1,}章", 4),
        (r"^\s*第[一二三四五六七八九十]{1,}节", 3),
        (r"^\s*[(（]?[一二三四五六七八九十]{1,}[)）]", 2),
        (r"^\s*[一二三四五六七八九十]{1,}\s*[、\.]", 2),
        (r"^\s*第[1234567890]{1,}章", 4),
        (r"^\s*第[1234567890]{1,}节", 3),
        (r"^\s*[(（]?[1234567890]{1,}[)）]", 5),
        (r"^\s*[1234567890]{1,}\s*[、\.]", 6),
        (r"^\s*(发行概况|发行人声明|重大事项提示|目\s*录|释\s*义|引\s*言)", 3)
    ]
    for pattern, mode in patterns:
        if re.search(pattern, title):
            return mode
    return None

def extract_table_title(char_list, rect, limit=None):
    """
    从指定的表格矩形上方limit范围内提取文本作为表格名称。已经弃用
    """
    if limit is None:
        title_region = LTRect(
                        bbox=(rect.x0, rect.y1 - 4, rect.x1, rect.y1 + 60), linewidth=1)  # 设定标题搜索区域
    else:
        title_region = LTRect(
                        bbox=(rect.x0, rect.y1 - 4, rect.x1, rect.y1 + limit), linewidth=1)  # 设定标题搜索区域
    text_list = filter_elements_within_rect(char_list, title_region)
    text_list = [item.get_text() for item in text_list]
    title_text = "".join(text_list)
    title_text = title_text.replace("适用", "").replace("不适用", "").replace("□", "").replace("√", "").replace("不", "")

    # 去除Excel不允许的字符(不能作为sheet标题)
    title_text = re.sub(r"(单\s*位)([\s：:]*)([个十百千万元吨公斤]+)", '', title_text)
    title_text = re.sub(r"(币\s*种)([\s：:]*)([人民币]+)", '', title_text)
    pattern = re.compile(r'[^\u4e00-\u9fa5a-zA-Z0-9()（）、,\.\s\-]')
    cleaned_title = re.sub(pattern, '', title_text)
    # 检查标题格式
    last_pattern, last_match, last_match_start = get_content_title_mode(cleaned_title,False)
    if last_pattern is not None:
        # 如果标题符合预设格式，返回清洗后的标题
        return cleaned_title[last_match_start:].strip()
    else:
        # 如果不符合预设格式，返回整个区域内的原始文本，同样需要清洗
        return cleaned_title
  
def group_rectangles(rectangles, mid_func ,threshold=2):
    """按横纵坐标对垂直线和水平线分组。

    Args:
        rectangles (_type_): 传入的rect列表
        mid_func (_type_): 垂直线用mid_x,水平线用mid_y
        threshold (int, optional): 判断是否划为一组的阈值。. Defaults to 2.

    Returns:
        _type_: _description_
    """
    # 按垂直中点分组
    groups = {}
    for rect in rectangles:
        my = mid_func(rect)
        placed = False
        for key in list(groups.keys()):
            if abs(my - key) < threshold:
                groups[key].append(rect)
                placed = True
                break
        if not placed:
            groups[my] = [rect]
    rect_list = []
    for key in groups.keys():
        rect_list += connect_rectangles(groups[key],mid_func)
    return rect_list

def connect_rectangles(rectangles,mid_func,distance_threshold=2):
    """将段连接起来的线条合并为一个线条

    Args:
        rectangles (_type_): 短线条列表
        mid_func (_type_): 垂直线用mid_x,水平线用mid_y
        distance_threshold (int, optional): 判断两个线条是否可以连接的阈值. Defaults to 2.

    Returns:
        _type_: 连接完后的线条列表。
    """
    if mid_func == mid_y:
        # 按x0排序矩形
        rectangles.sort(key=lambda rect: rect.bbox[0])

        # 存储连接后的矩形
        connected_rects = []

        # 初始化第一个矩形
        current_rect = rectangles[0]

        for rect in rectangles[1:]:
            # 检查当前矩形的x1是否足够接近下一个矩形的x0
            if current_rect.bbox[2] >= rect.bbox[0] - distance_threshold:  # 有重叠或紧挨着
                # 扩展当前矩形的x1到下一个矩形的x1
                current_rect = LTRect(
                    bbox=(current_rect.bbox[0], current_rect.bbox[1],
                        max(current_rect.bbox[2], rect.bbox[2]),
                        current_rect.bbox[3]),linewidth=1)
            else:
                # 没有重叠，保存当前矩形，并开始新的连接
                connected_rects.append(current_rect)
                current_rect = rect
        
        # 添加最后一个矩形
        connected_rects.append(current_rect)
    if mid_func == mid_x:
        # 按x0排序矩形
        rectangles.sort(key=lambda rect: rect.bbox[1])

        # 存储连接后的矩形
        connected_rects = []

        # 初始化第一个矩形
        current_rect = rectangles[0]

        for rect in rectangles[1:]:
            # 检查当前矩形的x1是否足够接近下一个矩形的x0
            if current_rect.bbox[3] >= rect.bbox[1] - distance_threshold:  # 有重叠或紧挨着
                # 扩展当前矩形的x1到下一个矩形的x1
                current_rect = LTRect(
                    bbox=(current_rect.bbox[0], current_rect.bbox[1],current_rect.bbox[2],
                        max(current_rect.bbox[3], rect.bbox[3])),linewidth=1)
            else:
                # 没有重叠，保存当前矩形，并开始新的连接
                connected_rects.append(current_rect)
                current_rect = rect
        
        # 添加最后一个矩形
        connected_rects.append(current_rect)
    return connected_rects

def copy_alignment(src_cell, dst_cell):
    """仅复制一个单元格的对齐属性，创建新的Alignment对象避免StyleProxy错误"""
    src_alignment = src_cell.alignment
    new_alignment = Alignment(horizontal=src_alignment.horizontal, 
                              vertical=src_alignment.vertical, 
                              text_rotation=src_alignment.text_rotation, 
                              wrap_text=src_alignment.wrap_text,
                              shrink_to_fit=src_alignment.shrink_to_fit,
                              indent=src_alignment.indent)
    dst_cell.alignment = new_alignment

def merge_sheet(sheet1,sheet2):
    """用于合并跨页表格

    Args:
        sheet1 (_type_): sheet1
        sheet2 (_type_): sheet2
    """
    # 计算开始插入新行的位置
    start_row = sheet1.max_row + 1
    # 遍历第二个工作表的每一行和每一列
    for i in range(1, sheet2.max_row + 1):
        for j in range(1, sheet2.max_column + 1):
            src_cell = sheet2.cell(row=i, column=j)
            dst_cell = sheet1.cell(row=start_row + i - 1, column=j)
            # 复制单元格的值和样式
            dst_cell.value = src_cell.value
            copy_alignment(src_cell, dst_cell)
    # 处理合并单元格
    for merge_cell in sheet2.merged_cells.ranges:
        new_merge = openpyxl.worksheet.cell_range.CellRange(
            min_row=merge_cell.min_row + start_row - 1,
            min_col=merge_cell.min_col,
            max_row=merge_cell.max_row + start_row - 1,
            max_col=merge_cell.max_col
        )
        sheet1.merge_cells(str(new_merge))

def get_box_line(box):
    """用于获取一个box对象的边框垂直线和边框水平线

    Args:
        box (_type_):box对象

    Returns:
        _type_: 垂直线，水平线。
    """
    ver_lines = []
    hor_lines = []
    ver_lines.append(LTRect(
                    bbox=(box.x0, box.y0, box.x0, box.y1),linewidth=1))
    ver_lines.append(LTRect(
                    bbox=(box.x1, box.y0, box.x1, box.y1),linewidth=1))
    hor_lines.append(LTRect(
                    bbox=(box.x0, box.y0, box.x1, box.y0),linewidth=1))
    hor_lines.append(LTRect(
                    bbox=(box.x0, box.y1, box.x1, box.y1),linewidth=1))
    return ver_lines,hor_lines

def sort_char_list(char_list):
    if char_list == []:
        return []
    # 初始化一个默认字典来存储分组后的结果
    grouped_dict = defaultdict(list)
    
    # 对char_list按y0排序
    sorted_list = sorted(char_list, key=lambda char: char.y0)
    
    # 初始化第一个分组的y0基准
    current_y0_group = sorted_list[0].y0
    
    for char in sorted_list:
        # 检查当前字符的y0是否在当前分组的范围内
        if abs(char.y0 - current_y0_group) <= 2:
            grouped_dict[current_y0_group].append(char)
        else:
            # 如果不在当前分组范围内，更新分组的y0基准
            current_y0_group = char.y0
            grouped_dict[current_y0_group].append(char)
    new_list = []
    char_dict = dict(grouped_dict)
    for key in sorted(char_dict.keys(),reverse=True):
            char_dict[key] = sorted(char_dict[key],key=lambda c :c.x0)
            new_list += char_dict[key]
    return new_list

def not_same_line(char1,char2):
    if abs(char1.y0-char2.y0)>char1.width*0.5:
        return True
    else:
        return False

def not_same_fontname(char1,char2):
    if is_chinese(char1.get_text()) and is_chinese(char2.get_text()):
        if char1.fontname != char2.fontname:
            return True
        else:
            return False
    else:
        return False

def is_chinese(char):
    # 判断一个字符是否是中文
    return re.match(r'[\u4e00-\u9fff]', char) is not None

def get_title_mode(title):
    """根据标题格式返回一个值，用于对不同格式的标题分级。
    想手动调整标题分级的情况可以调整返回的值。

    Args:
        title (str): 标题

    Returns:
        int : 代表格式的分类
    """
    pattern_list = [r"^第[一二三四五六七八九十]{1,}章",
                    r"^[(（]?[一二三四五六七八九十]{1,}[)）][、\.]?",
                    r"^[一二三四五六七八九十]{1,}\s*[、\.]",
                    r"^第[一二三四五六七八九十]{1,}节",
                    r"^第[1234567890]{1,}章",
                    r"^[(（]?[1234567890]{1,}[)）][、\.]?",
                    r"^[1234567890]{1,2}\s*[、\.]",
                    r"^[1234567890]{1,}\s+",
                    r"^第[1234567890]{1,}节",
                    r"^[abcdefghijkABCDEFGHIJK]\s*[\.、]"]
    
    for pattern in pattern_list:
        if re.search(pattern,title.strip()):
            return re.sub(pattern,"",title.strip()).strip()
    pattern_list = [r"表$",
                    r"说明[：:]?$",
                    r"[：:]$",
                    r"情况$",
                    r"在建产能$",
                    r"现有产能$",
                    r"的资金$"]
    for pattern in pattern_list:
        if re.search(pattern,title.strip()):
            if len(title)<30:
                return title.strip()
    return None

def find_table_title(text_table_name,text,ws):
    line = "\t"+text_table_name
    new_title = ws.title
    line_index = text.index(line)
    table_title = ""
    unit = ""
    try_num = 1
    while try_num<10:
        if text[line_index-try_num].startswith("\t"):
                break
        if get_title_mode(text[line_index-try_num]) is None:
            try_num += 1
        else :
            table_title = get_title_mode(text[line_index-try_num])
            new_title = re.sub(r"\s","",line).split("T")[0] + table_title
            new_title = re.sub(r'[\\\/\*\?\[\]\:\']',"",new_title)[:31]
            ws.title = new_title
            break
    try_num = 1
    while try_num<3:
        if text[line_index-try_num].startswith("\t"):
                break
        if not re.search(r"(单\s*位\s*[:：]\s*)([个十百千万亿元吨公斤%]+)",text[line_index-try_num]):
            try_num += 1
        else :
            unit = re.search(r"(单\s*位\s*[:：]\s*)([个十百千万亿元吨公斤%]+)",text[line_index-try_num]).group(2)
            ws.append([])
            ws.append(["单位:",unit])
            break
    return new_title[:31]

def group_by_y0(char_list):
    # 初始化一个默认字典来存储分组后的结果
    grouped_dict = defaultdict(list)
    if char_list == [] :
        return {}
    # 对char_list按y0排序
    sorted_list = sorted(char_list, key=lambda char: char.y0)
    
    # 初始化第一个分组的y0基准
    current_y0_group = sorted_list[0].y0
    
    for char in sorted_list:
        # 检查当前字符的y0是否在当前分组的范围内
        if abs(char.y0 - current_y0_group) <= 2:
            grouped_dict[current_y0_group].append(char)
        else:
            # 如果不在当前分组范围内，更新分组的y0基准
            current_y0_group = char.y0
            grouped_dict[current_y0_group].append(char)
    
    return dict(grouped_dict)

def process_text_rect(char_list,text_rect):
    text_lines = []
    line_rects = []
    temp_text= ""
    text_char_list = filter_elements_within_rect(char_list,text_rect)
    if text_char_list != []:
        char_dict = group_by_y0(text_char_list)
        for key in sorted(char_dict.keys(),reverse=True):
            char_dict[key] = sorted(char_dict[key],key=lambda c :c.x0)
            line_start_char = char_dict[key][0]
            line_end_char = char_dict[key][0]
            recent_char = char_dict[key][0]
            for text_char in char_dict[key]:
                if not_same_fontname(text_char,recent_char) :
                    text_lines.append(temp_text)
                    temp_text = ""
                    line_end_char = recent_char
                    line_rects.append(LTRect(
                            bbox=(line_start_char.x0,line_start_char.y0,line_end_char.x1,line_end_char.y1),linewidth=1) )
                    line_start_char = text_char
                elif abs(text_char.x1 - recent_char.x1)>1*(recent_char.width+text_char.width):
                    temp_text += " "
                temp_text+=text_char.get_text()
                recent_char = text_char
                line_end_char = recent_char
            text_lines.append(temp_text)
            line_rects.append(LTRect(
                                bbox=(line_start_char.x0,line_start_char.y0,line_end_char.x1,line_end_char.y1),linewidth=1) )
            temp_text = ""
    # for text_line in text_lines:
    #     print(text_line)
    return text_lines,line_rects

def  get_title_mode_line(title):
    """根据标题格式返回一个值，用于对不同格式的标题分级。
    想手动调整标题分级的情况可以调整返回的值。

    Args:
        title (str): 标题

    Returns:
        int : 代表格式的分类
    """
    pattern_list = [r"^第[一二三四五六七八九十]{1,}章",
                    r"^[(（]?[一二三四五六七八九十]{1,}[)）][、\.]?",
                    r"^[一二三四五六七八九十]{1,}\s*[、\.]",
                    r"^第[一二三四五六七八九十]{1,}节",
                    r"^第[1234567890]{1,}章",
                    r"^[(（]?[1234567890]{1,}[)）][、\.]?",
                    r"^[1234567890]{1,}\s*[、\.]",
                    r"^[1234567890]{1,}\s+",
                    r"^第[1234567890]{1,}节",
                    r"^[abcdefghijkABCDEFGHIJK]\s*[\.、]"]
    for pattern in pattern_list:
        if re.search(pattern,title.strip()):
            return re.sub(pattern,"",title.strip()).strip()
    return None

def process_page_lines(file_text_lines,file_line_rects,most_x0,most_x1):
    concated_result = []
    temp_paragraph = ""
    status = False
    if file_text_lines == []:
        return []
    recent_line_rect = file_line_rects[0]
    for line_index , file_line_rect in enumerate(file_line_rects):
        line_text = file_text_lines[line_index]
        if status:
            if (abs(file_line_rect.x0-most_x0) <3*(file_line_rect.height) 
            and not file_text_lines[line_index].startswith("\t") 
            and get_title_mode_line(file_text_lines[line_index]) is None
            and abs(recent_line_rect.y0 - file_line_rect.y0) < 3*(file_line_rect.height)):
                temp_paragraph += file_text_lines[line_index]
                if abs(most_x1-file_line_rect.x1)>3*(recent_line_rect.height):
                    concated_result.append(temp_paragraph)
                    temp_paragraph = ""
                    status = False
            else:
                concated_result.append(temp_paragraph)
                temp_paragraph = ""
                status = False
                if (abs(file_line_rect.x0-most_x0)<3*(file_line_rect.height) 
                    and abs(most_x1-file_line_rect.x1)<3*(file_line_rect.height)
                    and not file_text_lines[line_index].startswith("\t")):
                    status = True
                    temp_paragraph = file_text_lines[line_index]
                else:
                    concated_result.append(file_text_lines[line_index])
        else:
            if file_text_lines[line_index].startswith("\t"):
                concated_result.append(file_text_lines[line_index])
            else:
                if abs(file_line_rect.x0-most_x0)<3*(file_line_rect.height) and abs(most_x1-file_line_rect.x1)<3*(file_line_rect.height):
                    status = True
                    temp_paragraph = file_text_lines[line_index]
                else:
                    concated_result.append(file_text_lines[line_index])
        recent_line_rect = file_line_rect
    if status:
        concated_result.append(temp_paragraph)
    concated_result = [line for line in concated_result if line != ""]
    return concated_result

def extract_tables_from_pdf(pdf_path,txts_dir,excels_dir, page_number = None):

    """主要的调用的函数，通过pdfminer解析pdf中的线条识别表格以及表格的格式，将识别的表格全部存入对应名称下excel文件，可以指定pdf的某一页，默认处理全部页面.

    Args:
        pdf_path (): pdf路径
        page_number (_type_, optional): 指定pdf某一页。 Defaults to None.
    """
    # 打开PDF文件
    with open(pdf_path, 'rb') as file:
        parser = PDFParser(file)
        document = PDFDocument(parser)

        wb = Workbook()

        # 创建资源管理器和参数
        rsrcmgr = PDFResourceManager()
        laparams = LAParams()
        device = PDFPageAggregator(rsrcmgr, laparams=laparams)
        interpreter = PDFPageInterpreter(rsrcmgr, device)
        recent_table = None
        recent_layout = None
        # 读取特定页面
        for page_index, page in enumerate(tqdm(PDFPage.create_pages(document))):
            if page_number is None:
                pass
            else :
                if page_index != page_number-1:
                    continue
            interpreter.process_page(page)
            layout = device.get_result()
            vertical_list = []
            horizontal_list = []
            # 遍历页面中的对象
            for obj in layout:
                if isinstance(obj, LTRect):
                    if obj.width<1 and obj.height>2:
                        vertical_list.append(obj)
                    if obj.height<1 and obj.width>2:
                        horizontal_list.append(obj)
                if isinstance(obj, LTLine):
                    if obj.width<1 and obj.height>2:
                        vertical_list.append(obj)
                    if obj.height<1 and obj.width>2:
                        horizontal_list.append(obj)
            if len(vertical_list+horizontal_list) <3 :
                recent_table = None
                recent_layout = layout
                continue
            vertical_list = group_rectangles(vertical_list,mid_x)
            horizontal_list = group_rectangles(horizontal_list,mid_y)
            rectangles = find_table_rect(vertical_list,horizontal_list,4)
            if len(rectangles) == 0 :
                recent_table = None
                recent_layout = layout
                continue
            vertical_list = group_rectangles(vertical_list,mid_x)
            horizontal_list = group_rectangles(horizontal_list,mid_y)
            rectangles = sorted(rectangles, key=lambda rect: (-rect.y1, rect.x1))
            # # 以下注释代码用于检查转化出问题的文件，查看表格区域以及表格线条是否识别出错。
            # # 初始化绘图
            # fig, ax = plt.subplots()
            # for r in rectangles: # 测试表格范围定位是否准确
            #     x0, y0, x1, y1 = r.bbox  # 示例边界值
            #     # 计算矩形的宽度和高度
            #     width = x1 - x0
            #     height = y1 - y0
            #     rect = Rectangle((x0, y0), width, height, linewidth=1, edgecolor='b', facecolor='b')
            #     ax.add_patch(rect)
            # # 测试表格中线条识别情况
            # for v in vertical_list:
            #     mid = mid_x(v)
            #     line = Line2D([mid, mid], [v.y0, v.y1], linewidth=1, color='red')
            #     ax.add_line(line)
            # for h in horizontal_list:
            #     mid = mid_y(h)
            #     line = Line2D([h.x0, h.x1], [mid, mid], linewidth=1, color='red')
            #     ax.add_line(line)
            # # 设置图的边界
            # # 页面尺寸
            # page_width = page.mediabox[2]  # 页面的宽度
            # page_height = page.mediabox[3]  # 页面的高度
            # ax.set_xlim(0, page_width)
            # ax.set_ylim(0, page_height)
            # ax.set_aspect('equal')
            # plt.show()
            # 字体大小一般为8，9
            # 页眉边距一般为75
            footer_thresholds = 75
            header_thresholds = 75
            footer_thresholds = 75
            header_thresholds = 75
            page_height = page.mediabox[3] 
            page_width = page.mediabox[2] 
            char_list = [] 
            for element in layout:
                    if isinstance(element, LTTextContainer):
                        get_char_element(element,char_list)
            char_list = sort_char_list(char_list)
            footer_rect = LTRect(
                        bbox=(0,0,page_width,footer_thresholds),linewidth=1) 
            footer_char = filter_elements_within_rect(char_list,footer_rect)
            footer_list = [item.get_text() for item in footer_char if item.get_text() != " "]
            while len(footer_list) == 0 and footer_thresholds < 100:
                footer_thresholds += 5
                footer_rect = LTRect(
                        bbox=(0,0,page_width,footer_thresholds),linewidth=1) 
                footer_char = filter_elements_within_rect(char_list,footer_rect)
                footer_list = [item.get_text() for item in footer_char if item.get_text() != " "]
            footer_list = [item.get_text() for item in footer_char if item.get_text() != " "]
            while len(footer_list) == 0 and footer_thresholds < 100:
                footer_thresholds += 5
                footer_rect = LTRect(
                        bbox=(0,0,page_width,footer_thresholds),linewidth=1) 
                footer_char = filter_elements_within_rect(char_list,footer_rect)
                footer_list = [item.get_text() for item in footer_char if item.get_text() != " "]
            header_rect = LTRect(
                        bbox=(0,page_height-header_thresholds,page_width,page_height),linewidth=1) 
                        bbox=(0,page_height-header_thresholds,page_width,page_height),linewidth=1) 
            header_char = filter_elements_within_rect(char_list,header_rect)
            header_list = [item.get_text() for item in header_char if item.get_text() != " "]
            while len(header_list) == 0 and header_thresholds < 100:
                header_thresholds += 5
                header_rect = LTRect(
                        bbox=(0,0,page_width,header_thresholds),linewidth=1) 
                footer_char = filter_elements_within_rect(char_list,footer_rect)
                footer_list = [item.get_text() for item in footer_char if item.get_text() != " "]
            recent_r = None
            text_rects = []
            rectangles_edge_broaden = 7
            if rectangles == [] :
                text_rects.append(LTRect(
                        bbox=(0,footer_thresholds,page_width,page_height-header_thresholds),linewidth=1))
            else :
                for r_index,rectangle in enumerate(rectangles):
                    if r_index ==0 :
                        if rectangle.y1-rectangles_edge_broaden < page_height-header_thresholds:
                            text_rects.append(LTRect(
                                bbox=(0,rectangle.y1-rectangles_edge_broaden,page_width,page_height-header_thresholds),linewidth=1))
                        else:
                            text_rects.append(LTRect(
                                bbox=(0,rectangle.y1-rectangles_edge_broaden,page_width,rectangle.y1-rectangles_edge_broaden),linewidth=1))
                    if r_index == len(rectangles)-1:
                        if rectangle.y0+rectangles_edge_broaden > footer_thresholds:
                            text_rects.append(LTRect(
                                bbox=(0,footer_thresholds,page_width,rectangle.y0+rectangles_edge_broaden),linewidth=1))
                        else:
                            text_rects.append(LTRect(
                                bbox=(0,footer_thresholds,page_width,footer_thresholds),linewidth=1))
                    if r_index != len(rectangles)-1 :
                        text_rects.append(LTRect(
                                bbox=(0,rectangles[r_index+1].y1-rectangles_edge_broaden,page_width,rectangles[r_index].y0+rectangles_edge_broaden),linewidth=1))
            for text_rect_index,text_rect in enumerate(text_rects):
                text_lines,line_rects = process_text_rect(char_list,text_rect)
                page_text_lines += text_lines
                page_line_rects += line_rects
                if text_rect_index != len(text_rects)-1 :
                    page_text_lines += [f"\tP_{page_index+1} T_{text_rect_index+1}"]
                    page_line_rects.append(rectangles[text_rect_index])
            concat_result += process_page_lines(page_text_lines,page_line_rects,most_x0,most_x1)
            for table_index,r in enumerate(rectangles):
                if table_index == 0:
                    before_rect = LTRect(
                        bbox=(0,r.y1,page_width,page_height),linewidth=1) 
                    before_char = filter_elements_within_rect(char_list,before_rect)
                    before_list = [item.get_text() for item in before_char if item.get_text() != " "]
                    before_list = [item for item in before_list if item not in header_list]
                    if before_list != []:
                        recent_table = None
                if recent_r:
                    limit = recent_r.y0 - r.y1 + 8
                    if limit >=60 :
                        limit = None
                elif r.y1 > page_height-header_thresholds:
                    limit = 0
                elif r.y1+60 > page_height-header_thresholds :
                    limit = page_height-header_thresholds - r.y1
                else :
                    limit = None
                table_char_list = filter_elements_within_rect(char_list,r)
                table_vers = filter_elements_within_rect(vertical_list,r)
                table_vers = group_rectangles(table_vers,mid_x,2)
                table_hors = filter_elements_within_rect(horizontal_list,r)
                table_hors = group_rectangles(table_hors,mid_y,2)
                if table_hors==[] or table_vers == []:
                    recent_table = None
                    continue
                if table_index == 0 and before_list == [] and recent_table is not None: # 需要与上一个sheet合并。
                    diff = recent_table[-1].x0 - r.x0 # 确保两张表的列能够对齐
                    recent_y_keys = [round(key-diff) for key in recent_table[4].keys()]
                    rows,table_x0,table_x1 = init_rows(table_hors)
                    columns,table_y0,table_y1  = init_columns(table_vers)
                    if max(recent_y_keys) == max(columns.keys()) : # 与上一页表格能够对齐
                        for key in recent_y_keys:
                            if key not in columns.keys():
                                columns[key] = []
                        ws = wb.create_sheet(title=f"temp")
                        try:
                            init_table(ws,rows,columns,table_x0,table_y0,table_x1,table_y1,table_char_list)
                        except Exception as e:
                            logger.error(f"init_table error for {pdf_path} at page {page_index}", e)  
                        merge_sheet(wb[recent_table[2]],wb["temp"])
                        # print(f"merge_sheet {recent_table[2]}")
                        temp_sheet = wb["temp"]
                        wb.remove(temp_sheet)
                        if table_index != len(rectangles)-1 :  
                            recent_table = None
                    else: # 与上一页表格无法对齐，但两表中间没有任何信息可以直接使用上一表格的标题，但不合并。
                        ws = wb.create_sheet(title=f"Page {page_index+1} Table {table_index+1}")
                        title = f"P_{page_index+1} T_{table_index+1}"+recent_table[2].split("T_")[1][1:]
                        ws.title = title
                        recent_table = None
                        # print(f"Extracted Table Title for Page {page_index+1} Table {table_index+1}: '{title}'")
                        try:
                            init_table(ws,rows,columns,table_x0,table_y0,table_x1,table_y1,table_char_list)
                        except Exception as e:
                            logger.error(f"init_table error for {pdf_path} at page {page_index}", e)
                        if table_index != len(rectangles)-1 :  
                            recent_table = None
                else: # 不用合并表格时
                    rows,table_x0,table_x1 = init_rows(table_hors)
                    columns,table_y0,table_y1  = init_columns(table_vers)
                    ws = wb.create_sheet(title=f"Page {page_index+1} Table {table_index+1}")
                    title = extract_table_title(char_list, r, limit)
                    if not title:  # 如果标题为空，设置默认标题
                        title = f"P_{page_index+1} T_{table_index+1}"
                    else :
                        title =  f"P_{page_index+1} T_{table_index+1} " + title
                    title = title[:30] # 只支持30个字符以内
                    ws.title = title
                    # print(f"Extracted Table Title for Page {page_index+1} Table {table_index+1}: '{title}'")
                    try:
                        init_table(ws,rows,columns,table_x0,table_y0,table_x1,table_y1,table_char_list)
                    except Exception as e:
                        logger.error(f"init_table error for {pdf_path} at page {page_index}", e)
                    if table_index != len(rectangles)-1 :  
                        recent_table = None
                if table_index == len(rectangles)-1 :  # 如果时页面中的最后一个表格，判断是否可能是跨页表格，如果可能则存入recent_table。
                    after_rect = LTRect(
                        bbox=(0,0,page_width,r.y0),linewidth=1) 
                    after_char = filter_elements_within_rect(char_list,after_rect)
                    after_list = [item.get_text() for item in after_char if item.get_text() != " "]
                    after_list = [item for item in after_list if item not in footer_list]
                    if after_list == [] :
                        if recent_table == None :
                            recent_table = (page_index,table_index,title,rows,columns,r)
                        else:
                            recent_table = recent_table 
                    else:
                        recent_table = None
                recent_r = r
        try:
            default_sheet = wb["Sheet"]# 移除默认创建的 "Sheet"
            wb.remove(default_sheet)
            wb.save(f'/data/financial_report_baijiu/公司公告/tables/{os.path.basename(pdf_path).split(".pdf")[0]}.xlsx')
            logger.info(f'Saved file {os.path.basename(pdf_path).split(".pdf")[0]}')
        except Exception as e:
            logger.error(f'Fail to save {os.path.basename(pdf_path).split(".pdf")[0]}', e)
 
def process_pdf(pdf_path):
    extract_tables_from_pdf(pdf_path)


if __name__ == '__main__':
    directory_path = '/data/financial_report_baijiu/公司公告'

    # file_path = '/data/financial_report_baijiu/公司公告/202414_03212_公司公告/2018-08-02：贵州茅台2018年半年度报告.pdf'
    # extract_tables_from_pdf(file_path)

    pdf_files = []
    processed = set()
    time_outs = []

    for _,_, table in os.walk('/data/financial_report_baijiu/公司公告/tables'):
        for t in table:
            processed.add(t[:-5])

    # with open('/data/shenyuge/lingyue-data-process/pdflux/logs/log2024-04-25 12:19.log', 'r') as f:
    #     while line:= f.readline():
    #         re_expresssion = re.search(r"Saved file (.+)",line)
    #         if re_expresssion:
    #             time_outs += [re_expresssion.groups()[0]]

    for file in glob.glob(f'{directory_path}/**/*.pdf', recursive=True):
        # Check if the file has a PDF extension
        # if (file.endswith('年度报告.pdf') and file.split('/')[-1][:-4] not in processed) or (file in time_outs):
        if file.endswith('年度报告.pdf'):
            # Construct the full path to the PDF file
            pdf_files.append(file)
            # process_pdf(file)

    # num_threads = 8  # You can adjust this based on your system's capabilities

    # # # Process PDF files using multi-threading
    # with ProcessPoolExecutor(max_workers=num_threads) as executor:
    #     list(tqdm(executor.map(process_pdf, pdf_files), total=len(pdf_files)))
        # {executor.submit(process_pdf, pdf_file): pdf_file for pdf_file in pdf_files}

    # pdf_path = "/home/luzhenye/PythonProject/表格提取/2024-04-03：贵州茅台：贵州茅台2023年年度报告.pdf"
    pdf_path = "/home/luzhenye/data/pdf/周报/食品饮料行业双周报（20240520-20240602）：政策优化，关注预期改善.pdf"
    page_number = None # 要提取的页码
    extract_tables_from_pdf(pdf_path,txts_dir="/home/luzhenye/data/txt/周报",excels_dir="/home/luzhenye/data/excel/周报")

    # with open("/home/luzhenye/PythonProject/表格提取/logs/extract_table_log2024-06-11 12:52.log","r",encoding="utf-8") as file:
    #     lines = file.readlines()
    # erro_files =[]
    # for line in lines:
    #     match = re.search(r"(ERROR -\s*)(.*?)(提取表格出现问题)",line)
    #     erro_files.append(match.group(2))
    # print(len(erro_files))
    # num_threads = 8  # You can adjust this based on your system's capabilities

    # # # Process PDF files using multi-threading
    # with ProcessPoolExecutor(max_workers=num_threads) as executor:
    #     list(tqdm(executor.map(process_pdf, erro_files), total=len(erro_files)))


