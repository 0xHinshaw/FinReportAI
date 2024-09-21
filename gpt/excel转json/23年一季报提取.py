import math
import pandas as pd
import openpyxl
import difflib
import re
import difflib
from pathlib import Path
import json
import requests
import json
import ast
import os
import re
from tqdm import tqdm
import openpyxl
import pandas as pd
import numpy as np
from text_process import text2json

def get_response_gpt4(content):
    """ 获取gpt-4模型的回复

    Args:
        content (_type_): 给gpt-4的问题

    Returns:
        _type_: 模型的回答
    """
    # url = "https://gateway.ai.cloudflare.com/v1/05c43e30f91a115d8153715954fd70ee/lingyue-ai/openai/chat/completions"
    # url = "https://api.deepseek.com/chat/completions"
    url = "https://api.kwwai.top/v1/chat/completions"
    headers = {
        # "Authorization": "Bearer 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",
        # "Authorization": "Bearer sk-246e62fbd9cc4d12bd1cd65a5a532c06",
        "Authorization": "Bearer sk-QeiIJwcjqnhybuSeBbC0F27eEc0b42529a4410194b362bBb",
        "Content-Type": "application/json"
    }
    data = {
        # "model": "gpt-4-0613",
        # "model": "gpt-4o",
        # "model": "deepseek-chat",
        "model": "gpt-4o",
        "messages": [
            {
            "role": "user",
            "content": f"{content}"
            }
        ],
        "stream" : False
        }
        
    response = requests.post(url, json=data, headers=headers)
    # 假设 response.text 是一个字符串，内容是有效的JSON
    json_string = response.text
    # 将JSON字符串转换为字典
    data_json = json.loads(json_string)
    return data_json["choices"][0]["message"]["content"]


def get_data_json_from_gpt(table,company,table_name):
    prompt =  str(table) +"\n"+\
            "please read the table above and extract all data based on the given format: \n" +\
            "table name: \n" +\
            "indicator \t classification \t date \t value \t year-on-year \n"  +\
            "...  \t ... \t ... \t ... \t ...\n"  +\
            "...  \t ... \t ... \t ... \t ... \n"  +\
            "Please note that some row or column names need to be combined," +\
            "If classification doesn't exist, set value as '--'," +\
            "Do not add spaces in data cells,"+\
            "date can only be date information"+\
            "If year-on-year doesn't exist, set value as '--'," +"\n"+\
            "Please output table only and strictly follow the above format!"
    res = get_response_gpt4(prompt)
    return res

   


def save_data_from_excel(excel_path,save_json_path,save_txt_path,table_name):
    company = os.path.splitext(os.path.basename(excel_path))[0].split("：")[-2]
    file_name = os.path.splitext(os.path.basename(excel_path))[0].split("：")[-1]
    defualt_year = re.search(r"([0-9]{4})",file_name).group()
    match = re.search(r"([一二三]*?\s*季)|(半年)",file_name)
    if match:
        defualt_quarter = match.group()
    else:
        defualt_quarter = "全年"
    wb = openpyxl.load_workbook(excel_path)
    # 获取所有工作表的名称
    sheet_names = wb.sheetnames
    close_sheet = difflib.get_close_matches(table_name, sheet_names, n=3, cutoff=0.2)
    if close_sheet:
        df = pd.read_excel(excel_path, close_sheet[0])
        print(close_sheet)
    else:
        print(f"未能找到{table_name}")
        return
    if df.iloc[-1,0] == "单位:":
        table_unit = df.iloc[-1,1] 
    else :
        table_unit = "--"
    if table_unit == "--":
        markdown_str = str(df.to_markdown(index=False))
    else :
        markdown_str = str(df.iloc[:-2,:].to_markdown(index=False))
    df.fillna('', inplace=True)
    # print(close_sheet, df)
    df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
    # print(df)
    res= get_data_json_from_gpt(markdown_str,company,table_name)
    with open(save_txt_path,"w",encoding="utf-8") as file :
        file.write(res)
    print(res)
    # 指定在以下列中所有行都必须非空
    res = text2json(res,company,table_name,defualt_year,defualt_quarter,table_unit)
    with open(save_json_path,"w",encoding="utf-8") as file :
        json.dump(res, file,ensure_ascii=False, indent=4)

if __name__ == "__main__":
    excel_path = "/home/luzhenye/data/excel/tables/2023-04-26：贵州茅台：贵州茅台2023年第一季度报告.xlsx" 
    save_data_dir = "/home/luzhenye/data/贵州茅台报告表格数据提取脚本/23年一季报/data" 
    # 指定要提取的表格
    table_names = ["主要会计数据和财务指标","非经常性损益项目和金额","主要会计数据、财务指标发生变动的情况、原因","销售情况","经销商情况"]
    for table_name in table_names:
        save_json_path = os.path.join(save_data_dir,f"{table_name}.json")  
        save_txt_path =   os.path.join(save_data_dir,f"{table_name}.txt")  
        save_data_from_excel(excel_path,save_json_path,save_txt_path,table_name)

# prompt = str(markdown_str) +"\n"+\
#             "please read the table above and extract all data based on the given format: \n" +\
#             "table name: \n" +\
#             "indicator \t classification \t date \t value \t year-on-year \n"  +\
#             "...  \t ... \t ... \t ... \t ...\n"  +\
#             "...  \t ... \t ... \t ... \t ... \n"  +\
#             "Please note that some row or column names need to be combined," +\
#             "If classification doesn't exist, set value as '--'," +\
#             "Do not add spaces in data cells,"+\
#             "date can only be date information"+\
#             "If year-on-year doesn't exist, set value as '--'," +"\n"+\
#             "Please output table only and strictly follow the above format!"
# print(str(markdown_str))
# res = """
# """
# with open(f"/home/luzhenye/data/贵州茅台报告表格数据提取脚本/年度报告提取/data/{table_name}.txt","w",encoding="utf-8") as file :
#     file.write(res)
# excel_path = "/home/luzhenye/data/excel/tables/2023-04-26：贵州茅台：贵州茅台2023年第一季度报告.xlsx" 
# company = os.path.splitext(os.path.basename(excel_path))[0].split("：")[-2]
# file_name = os.path.splitext(os.path.basename(excel_path))[0].split("：")[-1]
# defualt_year = re.search(r"([0-9]{4})",file_name).group()
# match = re.search(r"([一二三]*?\s*季)|(半年)",file_name)
# if match:
#     defualt_quarter = match.group()
# else:
#     defualt_quarter = "全年"
# table_name = "主要会计数据、财务指标发生变动的情况、原因"
# wb = openpyxl.load_workbook(excel_path)
# # 获取所有工作表的名称
# sheet_names = wb.sheetnames
# close_sheet = difflib.get_close_matches(table_name, sheet_names, n=3, cutoff=0.2)
# if close_sheet:
#     df = pd.read_excel(excel_path, close_sheet[0])
#     print(close_sheet)
# else:
#     print(f"未能找到{table_name}")
# if df.iloc[-1,0] == "单位:":
#     table_unit = df.iloc[-1,1] 
# else :
#     table_unit = "--"
# with open(f"/home/luzhenye/data/贵州茅台报告表格数据提取脚本/23年一季报/data/{table_name}.txt","r",encoding="utf-8") as file:
#     text = file.read()
#     # print(text)
#     res = text2json(text,company,table_name,defualt_year,defualt_quarter,table_unit)
#     with open(f"/home/luzhenye/data/贵州茅台报告表格数据提取脚本/23年一季报/data/{table_name}.json","w",encoding="utf-8") as file :
#         json.dump(res, file,ensure_ascii=False, indent=4)

