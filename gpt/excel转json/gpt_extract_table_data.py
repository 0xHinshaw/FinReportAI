from datetime import datetime
from concurrent.futures import ProcessPoolExecutor
import math
import pandas as pd
import openpyxl
import difflib
import re
import difflib
from pathlib import Path
import json
import requests
import json
import ast
import os
import re
from tqdm import tqdm
import openpyxl
import pandas as pd
import numpy as np
import logging
from logging import Logger
from functools import partial
from text_process import text2json

def get_response_gpt4(content):
    """ 获取gpt-4模型的回复

    Args:
        content (_type_): 给gpt-4的问题

    Returns:
        _type_: 模型的回答
    """
    # url = "https://gateway.ai.cloudflare.com/v1/05c43e30f91a115d8153715954fd70ee/lingyue-ai/openai/chat/completions"
    # url = "https://api.deepseek.com/chat/completions"
    url = "https://api.kwwai.top/v1/chat/completions"
    headers = {
        # "Authorization": "Bearer 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",
        # "Authorization": "Bearer sk-246e62fbd9cc4d12bd1cd65a5a532c06",
        "Authorization": "Bearer sk-QeiIJwcjqnhybuSeBbC0F27eEc0b42529a4410194b362bBb",
        "Content-Type": "application/json"
    }
    data = {
        # "model": "gpt-4-0613",
        # "model": "gpt-4o",
        # "model": "deepseek-chat",
        "model": "gpt-4o",
        "messages": [
            {
            "role": "user",
            "content": f"{content}"
            }
        ],
        "stream" : False
        }
        
    response = requests.post(url, json=data, headers=headers)
    # 假设 response.text 是一个字符串，内容是有效的JSON
    json_string = response.text
    # 将JSON字符串转换为字典
    data_json = json.loads(json_string)
    return data_json["choices"][0]["message"]["content"]


def get_data_json_from_gpt(table,company,table_name):
    prompt =  str(table) +"\n"+\
            "please read the table above and extract all data based on the given format: \n" +\
            "table name: \n" +\
            "indicator \t classification \t date \t value \t year-on-year \n"  +\
            "...  \t ... \t ... \t ... \t ...\n"  +\
            "...  \t ... \t ... \t ... \t ... \n"  +\
            "Please note that some row or column names need to be combined," +\
            "If classification doesn't exist, set value as '--'," +\
            "Do not add spaces in data cells,"+\
            "date can only be date information"+\
            "If year-on-year doesn't exist, set value as '--'," +"\n"+\
            "Please output table only and strictly follow the above format!"
    res = get_response_gpt4(prompt)
    return res

   


def save_data_from_excel(table_name,excel_path,company,defualt_year,defualt_quarter,save_data_dir,log = False,logger = None):
    try:
        wb = openpyxl.load_workbook(excel_path)
        # 获取所有工作表的名称
        sheet_names = wb.sheetnames
        close_sheet = difflib.get_close_matches(table_name, sheet_names, n=3, cutoff=0.6)
        if close_sheet:
            df = pd.read_excel(excel_path, close_sheet[0])
            if log:
                logger.info(f"目标表名：{table_name} 实际找到的表: {close_sheet[0]}")
        else:
            if log:
                logger.error(f"{excel_path} 中未能找到 {table_name}")
            raise ValueError(f"{excel_path} 中未能找到 {table_name}")
        if df.iloc[-1,0] == "单位:":
            table_unit = df.iloc[-1,1] 
        else :
            table_unit = "--"
        if table_unit == "--":
            markdown_str = str(df.to_markdown(index=False))
        else :
            markdown_str = str(df.iloc[:-2,:].to_markdown(index=False))
        df.fillna('', inplace=True)
        # print(close_sheet, df)
        df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)
        # print(df)
        res= get_data_json_from_gpt(markdown_str,company,table_name)
        close_sheet[0] = re.sub(r"P_[0-9]{1,}","",close_sheet[0])
        save_json_path = os.path.join(save_data_dir,f"{close_sheet[0]}.json")  
        save_txt_path =   os.path.join(save_data_dir,f"{close_sheet[0]}.txt")
        with open(save_txt_path,"w",encoding="utf-8") as file :
            file.write(res)
        # print(res)
        # 指定在以下列中所有行都必须非空
        res = text2json(res,company,table_name,defualt_year,defualt_quarter,table_unit)
        with open(save_json_path,"w",encoding="utf-8") as file :
            json.dump(res, file,ensure_ascii=False, indent=4)
    except :
            if log:
                logger.error(f"{excel_path} {table_name} 提取出错")

def process_company_name(company):
    company = re.sub("＊","",company)
    company = re.sub("新疆伊力特实业股份有限公司","伊力特",company)
    company = re.sub("河北衡水老白干酒业股份有限公司","老白干酒",company)
    company = re.sub("安徽口子酒业股份有限公司","口子窖",company)
    company = re.sub("金徽酒股份有限公司","金徽酒",company)
    company = re.sub("江苏今世缘酒业股份有限公司","今世缘",company)
    company = re.sub("上海贵酒股份有限公司","岩石股份",company)
    company = re.sub("ST匹凸","岩石股份",company)
    company = re.sub("ST岩石","岩石股份",company)
    company = re.sub("ST皇台","皇台酒业",company)
    company = re.sub("ST舍得","舍得酒业",company)
    company = re.sub("青青稞酒","天佑德酒",company)


    return company

def process_excel(excel_path,table_names = None,select_company = None,select_year = None,select_quarter = None,log = False,logger = None,fix=False):
    file_name = os.path.splitext(os.path.basename(excel_path))[0].split("：")[-1]
    match = re.search(r"([:：])(.*?)([:：]?[0-9]{4})",excel_path)
    if match:
        company = re.sub(r"\s","",match.group(2))
        company = process_company_name(company)
        if re.search(r"[：:]",company):
            company = match.group(2).split("：")[0]
    else :
        if log:
            logger.error(f"{excel_path} 中未能识别到公司")
        return
    if select_company == None:
        pass
    elif company != select_company:
        return
    defualt_year = re.search(r"([0-9]{4})",file_name).group()
    match = re.search(r"([一二三]*?\s*季)|(半年)",file_name)
    if match:
        defualt_quarter = match.group()
    else:
        defualt_quarter = "全年"
    if select_year == None:
        pass
    elif defualt_year != select_year:
        return
    if select_quarter == None:
        pass
    elif defualt_quarter != select_quarter:
        return
    save_data_dir = f"/home/luzhenye/data/json/定期报告/{company}/{defualt_year}/{defualt_quarter}" 
    if fix:
        if not os.path.exists(save_data_dir):
            os.makedirs(save_data_dir)
    else:
        if not os.path.exists(save_data_dir):
            os.makedirs(save_data_dir)
        else:
            return
    # 指定要提取的表格
    if table_names == None:
        table_names = tables_for_quarter(defualt_quarter)
    if log:
        logger.info(f"开始处理{company}-{file_name}")
    save_data_from_excel_with_params = partial(save_data_from_excel, excel_path = excel_path,company=company,defualt_year=defualt_year,
                                               defualt_quarter=defualt_quarter,save_data_dir=save_data_dir,log = log,logger = logger)
    num_threads = 8  # You can adjust this based on your system's capabilities

    # # # Process PDF files using multi-threading
    with ProcessPoolExecutor(max_workers=num_threads) as executor:
        list(tqdm(executor.map(save_data_from_excel_with_params,table_names),total=len(table_names)))


def tables_for_quarter(quarter):
    # table_dict = {

    #     "一季": ["主要会计数据和财务指标", "非经常性损益项目和金额", "主要会计数据和财务指标发生变动的情况和原因"],
    #     "半年": ["主要会计数据","主要财务指标", "主要财务数据同比变动情况", "财务报表相关科目变动分析表", "非经常性损益项目及金额", 
    #             "资产构成重大变动情况", "销售费用情况", "资产及负债状况", "重大股权投资"],
    #     "三季": ["主要会计数据和财务指标", "非经常性损益项目级金额", "主要会计数据和财务指标发生变动的情况和原因"],
    #     "全年": ["主要会计数据","主要财务指标", "利润表及现金流量表相关科目变动分析表","非经常性损益项目及金额", "分季度主要财务指标", "财务报表相关科目变动分析表", "营业收入构成", 
    #             "公司主要销售客户情况","产销量情况分析表", "公司研发投入情况","主营业务分行业、分产品、分地区情况", "现金流",
    #               "资产构成重大变动情况", "费用","公司利润分配及资本公积金转增股本情况", "在建产能","现有产能","成本分析表","销售渠道","区域情况",
    #               "经销商情况","线上销售情况","按不同类型披露公司主营业务构成","资产及负债状况","采购金额"]

    # }
    table_dict = {
        "一季": ["主要会计数据","主要财务指标", "非经常性损益项目和金额","主要会计数据、财务指标发生变动的情况、原因"],
        "半年": ["主要财务指标", "主要会计数据", "财务报表相关科目变动分析表", "非经常性损益项目及金额", "按不同类型披露公司主营业务构成", 
         "按照区域分类披露经销商数量", "期末成品酒、半成品酒的库存", "资产及负债状况", "销售费用构成情况", "资产构成重大变动情况", 
         "以公允价值计量的资产和负债", "证券投资情况" ],
        "三季": ["主要会计数据和财务指标", "非经常性损益项目和金额","主要会计数据、财务指标发生变动的情况、原因" ],
        "全年": ["主要会计数据","主要财务指标", "非经常性损益项目及金额", "分季度主要财务指标", "按照区域分类披露经销商数量", 
         "财务报表相关科目变动分析表", "利润表及现金流量表相关科目变动分析表", "期末成品酒、半成品酒的库存", 
         "主营业务分行业、分产品、分地区、分销售模式情况", "产销量情况分析表", "销售费用构成情况", "成本分析表", "费用", "研发投入情况表", 
         "研发人员情况表", "现金流", "资产及负债状况", "现有产能", "在建产能", "产品情况", "按不同类型披露公司主营业务构成", 
         "资产构成重大变动情况", "本报告期利润分配及资本公积金转增股本预案", "采购金额", 
         "公司主要供应商情况", "证券投资情况", "销售渠道", "以公允价值计量的金融资产", "主要控股参股公司分析"]
    }
    return table_dict[quarter]

def get_all_company_of_file(excel_dir):
    all_company = []
    all_excel = os.listdir(excel_dir)
    for file in all_excel:
        file_name = os.path.splitext(file)[0]
        match = re.search(r"([:：])(.*?)([:：]?[0-9]{4})",file)
        if match:
            company = re.sub(r"\s","",match.group(2))
            company = process_company_name(company)
            if re.search(r"[：:]",company):
                company = match.group(2).split("：")[0]
            all_company.append(company)
        else:
            print(f"公司名字/年份识别失败：{file_name}")
    return list(set(all_company))
    
def init_logger(log_file_path = None):
    logger = logging.getLogger('my_logger')
    logger.setLevel(logging.DEBUG)
    if log_file_path is None:
        log_file_path=f'/home/luzhenye/PythonProject/gpt/excel转json/logs/通用提取{datetime.now().strftime("%Y-%m-%d %H:%M")}.log'
    file_handler = logging.FileHandler(log_file_path)
    # 创建并设置日志格式
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    file_handler.setFormatter(formatter)
    # 将文件处理器添加到 Logger
    logger.addHandler(file_handler)
    return logger

if __name__ == "__main__":
    log = False
    all_company=get_all_company_of_file("/home/luzhenye/data/excel/定期报告")
    print(all_company)
    # 启用logger
    # Logger = init_logger()
    # log = True
   

    # excel_dir = "/home/luzhenye/data/excel/定期报告"
    # excel_names = os.listdir(excel_dir)
    # for excel_name in excel_names:
    #     excel_path = os.path.join(excel_dir,excel_name) 
    #     table_names = None
    #     process_excel(excel_path,table_names,select_company="迎驾贡酒",log = log)
    


# prompt = str(markdown_str) +"\n"+\
#             "please read the table above and extract all data based on the given format: \n" +\
#             "table name: \n" +\
#             "indicator \t classification \t date \t value \t year-on-year \n"  +\
#             "...  \t ... \t ... \t ... \t ...\n"  +\
#             "...  \t ... \t ... \t ... \t ... \n"  +\
#             "Please note that some row or column names need to be combined," +\
#             "If classification doesn't exist, set value as '--'," +\
#             "Do not add spaces in data cells,"+\
#             "date can only be date information"+\
#             "If year-on-year doesn't exist, set value as '--'," +"\n"+\
#             "Please output table only and strictly follow the above format!"
# print(str(markdown_str))
# res = """
# """
# with open(f"/home/luzhenye/data/贵州茅台报告表格数据提取脚本/年度报告提取/data/{table_name}.txt","w",encoding="utf-8") as file :
#     file.write(res)
# excel_path = "/home/luzhenye/data/excel/tables/2023-04-26：贵州茅台：贵州茅台2023年第一季度报告.xlsx" 
# company = os.path.splitext(os.path.basename(excel_path))[0].split("：")[-2]
# file_name = os.path.splitext(os.path.basename(excel_path))[0].split("：")[-1]
# defualt_year = re.search(r"([0-9]{4})",file_name).group()
# match = re.search(r"([一二三]*?\s*季)|(半年)",file_name)
# if match:
#     defualt_quarter = match.group()
# else:
#     defualt_quarter = "全年"
# table_name = "主要会计数据、财务指标发生变动的情况、原因"
# wb = openpyxl.load_workbook(excel_path)
# # 获取所有工作表的名称
# sheet_names = wb.sheetnames
# close_sheet = difflib.get_close_matches(table_name, sheet_names, n=3, cutoff=0.2)
# if close_sheet:
#     df = pd.read_excel(excel_path, close_sheet[0])
#     print(close_sheet)
# else:
#     print(f"未能找到{table_name}")
# if df.iloc[-1,0] == "单位:":
#     table_unit = df.iloc[-1,1] 
# else :
#     table_unit = "--"
# with open(f"/home/luzhenye/data/贵州茅台报告表格数据提取脚本/23年一季报/data/{table_name}.txt","r",encoding="utf-8") as file:
#     text = file.read()
#     # print(text)
#     res = text2json(text,company,table_name,defualt_year,defualt_quarter,table_unit)
#     with open(f"/home/luzhenye/data/贵州茅台报告表格数据提取脚本/23年一季报/data/{table_name}.json","w",encoding="utf-8") as file :
#         json.dump(res, file,ensure_ascii=False, indent=4)

